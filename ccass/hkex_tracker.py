"""
CCASS Shareholding Tracker — Consolidated Module
Combines HKEX direct tracker, Webb-site historical tracker, and disk caching.

Classes:
    HKEXCCASSTracker      — Fetch CCASS data from HKEX (last 12 months)
    CachedHKEXTracker     — Adds file-based caching on top of HKEXCCASSTracker
    WebbCCASSTracker      — Fetch CCASS data from Webb-site (historical, beyond 12 months)
    HistoricalCCASSTracker — Orchestrator: multi-date snapshots + Excel report generation
"""

import hashlib
import json
import os
import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time


# ─── Utility ────────────────────────────────────────────────────────────────────

def timing_decorator(func):
    """Decorator to measure function execution time"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        duration = end_time - start_time
        if isinstance(result, dict):
            if 'timing' not in result:
                result['timing'] = {}
            result['timing'][func.__name__] = duration
        print(f"⏱️  {func.__name__}: {duration:.2f}s")
        return result
    return wrapper


# ─── 1. HKEX CCASS Tracker ──────────────────────────────────────────────────────

class HKEXCCASSTracker:
    """Track CCASS shareholding movements for HK stocks via HKEX (last 12 months)."""

    def __init__(self, enable_timing=True):
        self.base_url = "https://www3.hkexnews.hk/sdw/search/searchsdw.aspx"
        self.session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=10, pool_maxsize=10, max_retries=3
        )
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        self.enable_timing = enable_timing
        self.timings = {}
        self._viewstate_cache = None
        self._viewstate_cache_time = None
        self._cache_duration = 60  # seconds

    def format_stock_code(self, ticker: str) -> str:
        ticker = ticker.replace('.HK', '').replace('.hk', '')
        return ticker.zfill(5)

    def get_viewstate_and_cookies(self, use_cache=True) -> tuple:
        if use_cache and self._viewstate_cache is not None:
            cache_age = time.time() - (self._viewstate_cache_time or 0)
            if cache_age < self._cache_duration:
                if self.enable_timing:
                    print(f"⏱️  Using cached viewstate (age: {cache_age:.1f}s)")
                return self._viewstate_cache

        step_start = time.time()
        try:
            response = self.session.get(self.base_url, headers=self.headers, timeout=15)
            response.raise_for_status()
            parse_start = time.time()
            # Use lxml parser which is much faster than html.parser
            soup = BeautifulSoup(response.text, 'lxml')
            
            # Extract hidden form fields
            viewstate = soup.find('input', {'id': '__VIEWSTATE'})
            viewstategenerator = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})
            eventvalidation = soup.find('input', {'id': '__EVENTVALIDATION'})
            
            parse_time = time.time() - parse_start
            total_time = time.time() - step_start
            
            if self.enable_timing:
                print(f"⏱️  get_viewstate: {total_time:.2f}s (request: {total_time - parse_time:.2f}s, parse: {parse_time:.2f}s)")
            
            self.timings['get_initial_page'] = total_time - parse_time
            self.timings['parse_initial_page'] = parse_time
            
            result = {
                '__VIEWSTATE': viewstate['value'] if viewstate else '',
                '__VIEWSTATEGENERATOR': viewstategenerator['value'] if viewstategenerator else '',
                '__EVENTVALIDATION': eventvalidation['value'] if eventvalidation else ''
            }
            
            # Cache the result
            self._viewstate_cache = result
            self._viewstate_cache_time = time.time()
            
            return result
        except Exception as e:
            print(f"Error getting initial page: {e}")
            return None
    
    def fetch_ccass_data(self, ticker: str, date: Optional[str] = None) -> Dict:
        """
        Fetch CCASS shareholding data for a given ticker and date
        
        Args:
            ticker: Stock ticker (e.g., "01985" or "01985.HK")
            date: Date in format YYYY/MM/DD (default: today)
        
        Returns:
            Dictionary containing the parsed data
        """
        total_start = time.time()
        self.timings = {}
        
        # Format the ticker
        stock_code = self.format_stock_code(ticker)
        
        # Use today's date if not provided
        if date is None:
            date = datetime.now().strftime('%Y/%m/%d')
        
        print(f"Fetching CCASS data for stock {stock_code} on {date}...")
        
        # Get initial page data
        form_data = self.get_viewstate_and_cookies()
        if not form_data:
            return {"error": "Failed to get initial page data"}
        
        # Prepare POST data
        post_data = {
            '__EVENTTARGET': '',
            '__EVENTARGUMENT': '',
            '__VIEWSTATE': form_data['__VIEWSTATE'],
            '__VIEWSTATEGENERATOR': form_data['__VIEWSTATEGENERATOR'],
            '__EVENTVALIDATION': form_data['__EVENTVALIDATION'],
            'txtShareholdingDate': date,
            'txtStockCode': stock_code,
            'txtStockName': '',
            'txtParticipantID': '',
            'txtParticipantName': '',
            'btnSearch': 'Search'
        }
        
        try:
            # Submit the search
            post_start = time.time()
            response = self.session.post(
                self.base_url,
                data=post_data,
                headers=self.headers,
                timeout=15  # Reduced timeout
            )
            response.raise_for_status()
            
            post_time = time.time() - post_start
            self.timings['submit_search'] = post_time
            
            if self.enable_timing:
                print(f"⏱️  submit_search: {post_time:.2f}s")
            
            # Debug: Save the response
            # with open('debug_response.html', 'w', encoding='utf-8') as f:
            #     f.write(response.text)
            
            # Parse the response
            parse_start = time.time()
            result = self.parse_response(response.text, stock_code, date)
            parse_time = time.time() - parse_start
            
            self.timings['parse_response'] = parse_time
            
            if self.enable_timing:
                print(f"⏱️  parse_response: {parse_time:.2f}s")
            
            total_time = time.time() - total_start
            self.timings['total_fetch'] = total_time
            result['timing'] = self.timings.copy()
            
            if self.enable_timing:
                print(f"⏱️  TOTAL: {total_time:.2f}s")
                self._print_timing_summary()
            
            return result
            
        except requests.exceptions.RequestException as e:
            return {"error": f"Request failed: {e}"}
    
    def parse_response(self, html: str, stock_code: str, date: str) -> Dict:
        """
        Parse the HTML response to extract CCASS shareholding data
        """
        # Use lxml parser which is faster
        soup = BeautifulSoup(html, 'lxml')
        
        result = {
            'stock_code': stock_code,
            'date': date,
            'stock_name': '',
            'total_issued_shares': '',
            'participants': [],
            'summary': {}
        }
        
        # Extract stock name from input field (it gets populated after search)
        stock_name_input = soup.find('input', {'id': 'txtStockName'})
        if stock_name_input and stock_name_input.get('value'):
            result['stock_name'] = stock_name_input['value']
        
        # Extract summary information
        summary_divs = soup.find_all('div', class_='ccass-search-summary-table')
        if summary_divs:
            for summary_div in summary_divs:
                category = summary_div.find('div', class_='summary-category')
                num_participants = summary_div.find('div', class_='number-of-participants')
                percent = summary_div.find('div', class_='percent-of-participants')
                
                if category:
                    cat_name = category.get_text(strip=True)
                    result['summary'][cat_name] = {
                        'participants': num_participants.find('div', class_='value').get_text(strip=True) if num_participants and num_participants.find('div', class_='value') else 'N/A',
                        'percentage': percent.find('div', class_='value').get_text(strip=True) if percent and percent.find('div', class_='value') else 'N/A'
                    }
        
        # Extract the main data table
        table = soup.find('table', class_='table-mobile-list')
        
        if table:
            # Parse data rows (skip header row)
            rows = table.find_all('tr')
            
            for row in rows[1:]:  # Skip header
                cols = row.find_all('td')
                
                if len(cols) >= 4:
                    # Extract data from mobile-list-body divs
                    participant_id_div = cols[0].find('div', class_='mobile-list-body')
                    participant_name_div = cols[1].find('div', class_='mobile-list-body')
                    # Note: col[2] is address, we skip it
                    shareholding_div = cols[3].find('div', class_='mobile-list-body') if len(cols) > 3 else None
                    percentage_div = cols[4].find('div', class_='mobile-list-body') if len(cols) > 4 else None
                    
                    if participant_id_div:
                        participant_data = {
                            'participant_id': participant_id_div.get_text(strip=True),
                            'participant_name': participant_name_div.get_text(strip=True) if participant_name_div else '',
                            'address': cols[2].find('div', class_='mobile-list-body').get_text(strip=True) if len(cols) > 2 and cols[2].find('div', class_='mobile-list-body') else '',
                            'shareholding': shareholding_div.get_text(strip=True) if shareholding_div else '',
                            'percentage': percentage_div.get_text(strip=True) if percentage_div else ''
                        }
                        result['participants'].append(participant_data)
        
        return result
    
    def display_results(self, data: Dict):
        """
        Display the CCASS data in a formatted table
        """
        if 'error' in data:
            print(f"\n❌ Error: {data['error']}")
            return
        
        print(f"\n{'='*80}")
        print(f"CCASS SHAREHOLDING DATA")
        print(f"{'='*80}")
        print(f"Stock Code: {data['stock_code']}")
        print(f"Stock Name: {data['stock_name']}")
        print(f"Date: {data['date']}")
        print(f"{'='*80}\n")
        
        # Display summary if available
        if data.get('summary'):
            print("SUMMARY:")
            print("-" * 80)
            for category, info in data['summary'].items():
                print(f"{category}:")
                print(f"  Number of Participants: {info.get('participants', 'N/A')}")
                print(f"  Percentage: {info.get('percentage', 'N/A')}")
            print(f"{'='*80}\n")
        
        if data['participants']:
            # Create DataFrame for better display
            df = pd.DataFrame(data['participants'])
            
            # Rename columns for better display
            df_display = df[['participant_id', 'participant_name', 'shareholding', 'percentage']].copy()
            df_display.columns = ['Participant ID', 'Participant Name', 'Shareholding', '% of Total']
            
            print("PARTICIPANT DETAILS:")
            print("-" * 80)
            
            # Set pandas display options for better formatting
            pd.set_option('display.max_colwidth', 40)
            pd.set_option('display.width', 120)
            
            print(df_display.to_string(index=False))
            print(f"\n{'='*80}")
            print(f"Total Participants: {len(data['participants'])}")
            
            # Calculate total shareholding if possible
            try:
                total_shares = sum([int(p['shareholding'].replace(',', '')) for p in data['participants']])
                print(f"Total Shares Held: {total_shares:,}")
            except:
                pass
            
            print(f"{'='*80}\n")
        else:
            print("No participant data found.")
            print(f"{'='*80}\n")
    
    def export_to_csv(self, data: Dict, filename: Optional[str] = None):
        """
        Export the data to CSV file
        """
        if 'error' in data or not data['participants']:
            print("No data to export.")
            return
        
        if filename is None:
            filename = f"CCASS_{data['stock_code']}_{data['date'].replace('/', '')}.csv"
        
        df = pd.DataFrame(data['participants'])
        df.to_csv(filename, index=False)
        print(f"✅ Data exported to {filename}")
    
    def _print_timing_summary(self):
        """Print a summary of timing information"""
        print(f"\n{'='*80}")
        print("PERFORMANCE BREAKDOWN")
        print(f"{'='*80}")
        
        timings_sorted = sorted(self.timings.items(), key=lambda x: x[1], reverse=True)
        
        total = self.timings.get('total_fetch', sum(self.timings.values()))
        
        for name, duration in timings_sorted:
            if name != 'total_fetch':
                percentage = (duration / total * 100) if total > 0 else 0
                bar_length = int(percentage / 2)
                bar = '█' * bar_length
                print(f"{name:.<30} {duration:>6.2f}s {bar:20} {percentage:>5.1f}%")
        
        print(f"{'='*80}\n")


# ─── 2. Cached HKEX Tracker ─────────────────────────────────────────────────────

class CachedHKEXTracker(HKEXCCASSTracker):
    """Enhanced HKEX tracker with file-based caching."""

    def __init__(self, cache_dir=".cache", cache_days=7, enable_timing=True):
        super().__init__(enable_timing=enable_timing)
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        self.cache_days = cache_days

    def _get_cache_key(self, stock_code: str, date: str) -> str:
        return hashlib.md5(f"{stock_code}_{date}".encode()).hexdigest()

    def _get_cache_path(self, cache_key: str) -> Path:
        return self.cache_dir / f"{cache_key}.json"

    def _is_cache_valid(self, cache_path: Path) -> bool:
        if not cache_path.exists():
            return False
        mtime = datetime.fromtimestamp(cache_path.stat().st_mtime)
        return (datetime.now() - mtime) < timedelta(days=self.cache_days)

    def _load_from_cache(self, cache_key: str) -> dict:
        cache_path = self._get_cache_path(cache_key)
        if self._is_cache_valid(cache_path):
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if self.enable_timing:
                        print(f"📦 Loaded from cache: {cache_path.name}")
                    return data
            except Exception as e:
                print(f"Warning: Failed to load cache: {e}")
        return None

    def _save_to_cache(self, cache_key: str, data: dict):
        cache_path = self._get_cache_path(cache_key)
        try:
            cache_data = {k: v for k, v in data.items() if k != 'timing'}
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            if self.enable_timing:
                print(f"💾 Saved to cache: {cache_path.name}")
        except Exception as e:
            print(f"Warning: Failed to save cache: {e}")

    def fetch_ccass_data(self, ticker: str, date: str = None, use_cache=True):
        stock_code = self.format_stock_code(ticker)
        if date is None:
            date = datetime.now().strftime('%Y/%m/%d')
        if use_cache:
            cache_key = self._get_cache_key(stock_code, date)
            cached_data = self._load_from_cache(cache_key)
            if cached_data:
                return cached_data
        data = super().fetch_ccass_data(ticker, date)
        if use_cache and 'error' not in data:
            cache_key = self._get_cache_key(stock_code, date)
            self._save_to_cache(cache_key, data)
        return data

    def clear_cache(self, older_than_days=None):
        cleared = 0
        for cache_file in self.cache_dir.glob("*.json"):
            should_delete = older_than_days is None
            if not should_delete:
                mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
                if (datetime.now() - mtime) > timedelta(days=older_than_days):
                    should_delete = True
            if should_delete:
                cache_file.unlink()
                cleared += 1
        print(f"🗑️  Cleared {cleared} cache files")
        return cleared

    def get_cache_stats(self):
        cache_files = list(self.cache_dir.glob("*.json"))
        if not cache_files:
            print("📊 Cache is empty")
            return
        total_size = sum(f.stat().st_size for f in cache_files)
        print(f"{'='*80}")
        print("CACHE STATISTICS")
        print(f"{'='*80}")
        print(f"Total files: {len(cache_files)}")
        print(f"Total size: {total_size / 1024:.2f} KB")
        print(f"Average size: {total_size / len(cache_files) / 1024:.2f} KB per file")
        print(f"Cache directory: {self.cache_dir.absolute()}")
        print(f"{'='*80}\n")


# ─── 3. Webb-site CCASS Tracker ─────────────────────────────────────────────────

class WebbCCASSTracker:
    """Fetch CCASS data from Webb-site — historical data beyond HKEX's 12-month limit."""

    def __init__(self, cache_dir: str = ".cache_webb", base_url_template: str = None):
        self.cache_dir = cache_dir
        self.base_url_template = base_url_template
        os.makedirs(cache_dir, exist_ok=True)

        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=1, pool_maxsize=1, max_retries=0
        )
        self.session.mount('https://', adapter)
        self.session.mount('http://', adapter)

    def _get_cache_path(self, stock_code: str, date_str: str) -> str:
        cache_key = f"{stock_code}_{date_str}"
        hash_key = hashlib.md5(cache_key.encode()).hexdigest()
        return os.path.join(self.cache_dir, f"{hash_key}.json")

    def _load_from_cache(self, cache_path: str) -> Optional[Dict]:
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return None
        return None

    def _save_to_cache(self, cache_path: str, data: Dict):
        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"⚠️  Cache save failed: {e}")

    def fetch_ccass_data(self, stock_code: str, date_str: str) -> dict:
        """Fetch CCASS participant holdings from Webb-site."""
        date_webb = date_str.replace('/', '-')
        cache_path = self._get_cache_path(stock_code, date_str)
        cached_data = self._load_from_cache(cache_path)
        if cached_data:
            return cached_data

        try:
            if self.base_url_template:
                url = re.sub(r'd=\d{4}-\d{2}-\d{2}', f'd={date_webb}', self.base_url_template)
            else:
                sc = str(int(stock_code)) if stock_code.isdigit() else stock_code
                url = f"https://webbsite.0xmd.com/ccass/choldings.asp?sc={sc}&d={date_webb}"

            headers = {'Referer': 'https://webbsite.0xmd.com/ccass/'}
            response = self.session.get(url, headers=headers, timeout=20, allow_redirects=True)

            if response.status_code != 200:
                return {'date': date_str, 'stock_code': stock_code, 'participants': [],
                        'error': f'HTTP {response.status_code}'}

            soup = BeautifulSoup(response.content, 'html.parser')
            participants = []
            table = soup.find('table', class_='optable yscroll') or soup.find('table', class_='optable')

            if table:
                rows = table.find_all('tr')
                for row in rows[1:]:
                    cells = row.find_all('td')
                    if len(cells) >= 6:
                        try:
                            participant_id = cells[1].get_text(strip=True)
                            participant_name = cells[2].get_text(strip=True)
                            shareholding_str = cells[3].get_text(strip=True).replace(',', '')
                            percentage_str = cells[5].get_text(strip=True)

                            if not participant_id or not participant_id[0].isalpha():
                                continue

                            try:
                                shareholding = int(shareholding_str) if shareholding_str.replace('-', '').replace('.', '').isdigit() else 0
                                percentage = float(percentage_str) if percentage_str and percentage_str != '-' else 0.0
                            except Exception:
                                shareholding = 0
                                percentage = 0.0

                            if participant_id and shareholding > 0:
                                participants.append({
                                    'participant_id': participant_id,
                                    'participant_name': participant_name,
                                    'shareholding': shareholding,
                                    'percentage': percentage,
                                })
                        except Exception:
                            continue

            data = {'date': date_str, 'stock_code': stock_code, 'participants': participants}
            self._save_to_cache(cache_path, data)
            return data

        except Exception as e:
            print(f"⚠️  Fetch error: {e}")
            return {'date': date_str, 'stock_code': stock_code, 'participants': [], 'error': str(e)}


# ─── 4. Historical CCASS Tracker (orchestrator) ─────────────────────────────────

class HistoricalCCASSTracker:
    """
    Multi-date snapshot orchestrator with Excel report generation.
    Works with either HKEX or Webb-site as data source.
    """

    def __init__(self, source: str = 'webb', base_url_template: str = None):
        """
        Args:
            source: 'hkex' or 'webb'
            base_url_template: Optional Webb-site URL template
        """
        if source == 'hkex':
            self.tracker = CachedHKEXTracker(enable_timing=False)
        else:
            self.tracker = WebbCCASSTracker(base_url_template=base_url_template)
        self.source = source

    def generate_snapshot_dates(
        self, start_date: str, end_date: str = None, interval_days: int = 30
    ) -> List[str]:
        start = datetime.strptime(start_date, "%Y/%m/%d")
        end = datetime.strptime(end_date, "%Y/%m/%d") if end_date else datetime.now()
        dates = []
        current = start
        while current <= end:
            dates.append(current.strftime("%Y/%m/%d"))
            current += timedelta(days=interval_days)
        if dates[-1] != end.strftime("%Y/%m/%d"):
            dates.append(end.strftime("%Y/%m/%d"))
        return dates

    def fetch_historical_snapshots(
        self, stock_code: str, snapshot_dates: List[str], delay_seconds: float = 1.0
    ) -> Dict[str, Dict]:
        snapshots = {}
        total = len(snapshot_dates)
        no_data_count = 0
        error_count = 0

        src_label = "HKEX" if self.source == 'hkex' else "Webb-site"
        print(f"\n📊 Fetching {total} snapshots for {stock_code} from {src_label}...")
        print(f"   From: {snapshot_dates[0]}")
        print(f"   To: {snapshot_dates[-1]}\n")

        consecutive_errors = 0

        for i, date in enumerate(snapshot_dates, 1):
            max_retries = 3
            retry_delay = delay_seconds

            for attempt in range(max_retries):
                try:
                    if attempt == 0:
                        print(f"   [{i}/{total}] {date}...", end=" ", flush=True)
                    else:
                        print(f"retry {attempt}...", end=" ", flush=True)

                    if self.source == 'hkex':
                        data = self.tracker.fetch_ccass_data(stock_code, date)
                    else:
                        data = self.tracker.fetch_ccass_data(stock_code, date)

                    if data and 'participants' in data and len(data['participants']) > 0:
                        snapshots[date] = data
                        print(f"✓ ({len(data['participants'])} participants)")
                        consecutive_errors = 0
                        break
                    elif 'error' in data:
                        if 'HTTP 403' in str(data.get('error', '')):
                            if attempt < max_retries - 1:
                                backoff_time = retry_delay * (2 ** attempt) * 2
                                time.sleep(backoff_time)
                                continue
                        error_count += 1
                        consecutive_errors += 1
                        print(f"✗ {data['error']}")
                        break
                    else:
                        no_data_count += 1
                        print("○ No data")
                        consecutive_errors = 0
                        break
                except Exception as e:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (2 ** attempt))
                        continue
                    else:
                        error_count += 1
                        consecutive_errors += 1
                        print(f"✗ Error: {e}")

            if consecutive_errors >= 3:
                delay_seconds = min(delay_seconds * 1.5, 10)
                print(f"   ⚠️  Increasing delay to {delay_seconds:.1f}s due to errors")
                consecutive_errors = 0

            if i < total:
                time.sleep(delay_seconds)

        print(f"\n   Summary: {len(snapshots)} successful, {no_data_count} no data, {error_count} errors\n")
        return snapshots

    # ── Report helpers ──────────────────────────────────────────────────────────

    def _calculate_holding_changes(self, snapshots: Dict[str, Dict]) -> pd.DataFrame:
        dates = sorted(snapshots.keys())
        if len(dates) < 2:
            return pd.DataFrame()

        recent_dates = dates[-7:] if len(dates) >= 7 else dates
        all_changes = []

        for i in range(1, len(recent_dates)):
            prev_date = recent_dates[i - 1]
            curr_date = recent_dates[i]
            prev_data = {p['participant_id']: p for p in snapshots[prev_date]['participants']}
            curr_data = {p['participant_id']: p for p in snapshots[curr_date]['participants']}
            all_pids = set(prev_data.keys()) | set(curr_data.keys())

            for pid in all_pids:
                curr_shares = curr_data[pid]['shareholding'] if pid in curr_data else 0
                prev_shares = prev_data[pid]['shareholding'] if pid in prev_data else 0
                change = curr_shares - prev_shares
                if change != 0:
                    curr_pct = curr_data[pid]['percentage'] if pid in curr_data else 0
                    pname = (curr_data[pid].get('participant_name', pid) if pid in curr_data
                             else prev_data[pid].get('participant_name', pid))
                    all_changes.append({
                        'Date': curr_date, 'From': prev_date, 'CCASS ID': pid,
                        'Name': pname, 'Change': change,
                        'Current Holding': curr_shares, 'Stake%': curr_pct,
                        '_abs_change': abs(change), '_date_sort': curr_date,
                    })

        df = pd.DataFrame(all_changes)
        if not df.empty:
            df = df.sort_values(['_date_sort', '_abs_change'], ascending=[False, False])
            df = df.drop(['_abs_change', '_date_sort'], axis=1)
            df['Change'] = df['Change'].apply(lambda x: f"{x:+,}")
            df['Current Holding'] = df['Current Holding'].apply(lambda x: f"{x:,}")
            df['Stake%'] = df['Stake%'].apply(lambda x: f"{x:.2f}")
            df.insert(0, 'Row', range(1, len(df) + 1))
        return df

    def _create_ccass_holdings_sheet(self, snapshots: Dict[str, Dict]) -> Dict:
        latest_date = max(snapshots.keys())
        participants = snapshots[latest_date]['participants']
        custodians = brokers = other = 0
        for p in participants:
            shares = p['shareholding']
            pid = p['participant_id']
            if pid.startswith('C'):
                custodians += shares
            elif pid.startswith('B'):
                brokers += shares
            else:
                other += shares
        total_ccass = sum(p['shareholding'] for p in participants)
        intermediaries = custodians + brokers + other
        total_issued = 0
        for p in participants:
            if p['percentage'] > 0:
                total_issued = int(p['shareholding'] / (p['percentage'] / 100))
                break
        if total_issued == 0:
            total_issued = int(total_ccass / 0.55)
        return {
            'custodians': custodians, 'brokers': brokers, 'other': other,
            'intermediaries': intermediaries, 'total_ccass': total_ccass,
            'total_issued': total_issued, 'participants': participants,
            'date': latest_date,
        }

    def _create_movement_timeline(
        self, snapshots: Dict[str, Dict], top_n: int = 5, movement_type: str = 'buyer'
    ) -> Tuple[List, List]:
        dates = sorted(snapshots.keys())
        participant_changes: Dict[str, dict] = {}

        for i in range(1, len(dates)):
            prev_data = {p['participant_id']: p for p in snapshots[dates[i - 1]]['participants']}
            curr_data = {p['participant_id']: p for p in snapshots[dates[i]]['participants']}
            all_pids = set(prev_data.keys()) | set(curr_data.keys())
            for pid in all_pids:
                if pid not in participant_changes:
                    src = curr_data.get(pid) or prev_data.get(pid)
                    participant_changes[pid] = {
                        'name': src['participant_name'] if src else pid,
                        'changes': {}, 'stakes': {}, 'holdings': {},
                        'total_change': 0, 'total_net_change': 0,
                    }
                curr_shares = curr_data[pid]['shareholding'] if pid in curr_data else 0
                prev_shares = prev_data[pid]['shareholding'] if pid in prev_data else 0
                change = curr_shares - prev_shares
                curr_pct = curr_data[pid]['percentage'] if pid in curr_data else 0
                participant_changes[pid]['changes'][dates[i]] = change
                participant_changes[pid]['stakes'][dates[i]] = curr_pct
                participant_changes[pid]['holdings'][dates[i]] = curr_shares
                participant_changes[pid]['total_change'] += abs(change)
                participant_changes[pid]['total_net_change'] += change

        if movement_type == 'buyer':
            candidates = []
            for pid, data in participant_changes.items():
                total_buys = sum(c for c in data['changes'].values() if c > 0)
                if total_buys > 0:
                    data['total_buys'] = total_buys
                    candidates.append((pid, data))
            top_participants = sorted(candidates, key=lambda x: x[1]['total_buys'], reverse=True)[:top_n]
        else:
            candidates = []
            for pid, data in participant_changes.items():
                total_sells = sum(abs(c) for c in data['changes'].values() if c < 0)
                if total_sells > 0:
                    data['total_sells'] = total_sells
                    candidates.append((pid, data))
            top_participants = sorted(candidates, key=lambda x: x[1]['total_sells'], reverse=True)[:top_n]

        return top_participants, dates

    # ── Excel report ────────────────────────────────────────────────────────────

    def generate_excel_report(
        self, stock_code: str, snapshots: Dict[str, Dict],
        start_date: str, output_dir: str = "reports"
    ) -> str:
        os.makedirs(output_dir, exist_ok=True)
        end_date = max(snapshots.keys()).replace('/', '-')
        src_tag = "HKEX" if self.source == 'hkex' else "Webb"
        filename = f"{stock_code}_{src_tag}_Historical_Analysis_{start_date.replace('/', '-')}_to_{end_date}.xlsx"
        filepath = os.path.join(output_dir, filename)

        print(f"\n📊 Generating Excel report...")
        wb = Workbook()
        wb.remove(wb.active)

        # ── Sheet 1: CCASS Holding Changes ──
        print("   Creating CCASS Holding Changes sheet...")
        ws = wb.create_sheet("CCASS Holding Changes")
        changes_df = self._calculate_holding_changes(snapshots)
        if not changes_df.empty:
            ws.append([f"CCASS Holding Changes for {stock_code}"])
            ws.merge_cells('A1:H1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.append([])

            current_date = None
            row_num = 3
            for _, row_data in changes_df.iterrows():
                date_pair = f"{row_data['From']} → {row_data['Date']}"
                if date_pair != current_date:
                    if current_date is not None:
                        ws.append([])
                        row_num += 1
                    ws.append([date_pair])
                    ws.merge_cells(f'A{row_num}:H{row_num}')
                    date_cell = ws[f'A{row_num}']
                    date_cell.font = Font(size=11, bold=True, color='FFFFFF')
                    date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    date_cell.alignment = Alignment(horizontal='center')
                    row_num += 1
                    ws.append(['#', 'CCASS ID', 'Name', 'Change', 'Current Holding', 'Stake%'])
                    for cell in ws[row_num]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    row_num += 1
                    current_date = date_pair
                    item_num = 1

                ws.append([
                    item_num, row_data['CCASS ID'], row_data['Name'],
                    row_data['Change'], row_data['Current Holding'], row_data['Stake%'],
                ])
                change_str = str(row_data['Change'])
                if change_str.startswith('+'):
                    ws.cell(row=row_num, column=4).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    ws.cell(row=row_num, column=4).font = Font(color='006100', bold=True)
                elif change_str.startswith('-'):
                    ws.cell(row=row_num, column=4).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    ws.cell(row=row_num, column=4).font = Font(color='9C0006', bold=True)
                for col_idx in range(1, 7):
                    ws.cell(row=row_num, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                row_num += 1
                item_num += 1

            for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], 1):
                max_length = max(
                    (len(str(cell.value)) for row in ws.iter_rows(min_col=col_idx, max_col=col_idx) for cell in row if cell.value),
                    default=0,
                )
                if col_idx == 3:
                    ws.column_dimensions[col_letter].width = max(45, min(max_length + 2, 60))
                elif col_idx == 1:
                    ws.column_dimensions[col_letter].width = max(6, max_length + 2)
                elif col_idx == 2:
                    ws.column_dimensions[col_letter].width = max(12, max_length + 2)
                else:
                    ws.column_dimensions[col_letter].width = max(15, max_length + 2)

        # ── Sheet 2: CCASS Holdings ──
        print("   Creating CCASS Holdings sheet...")
        ws = wb.create_sheet("CCASS Holdings")
        hd = self._create_ccass_holdings_sheet(snapshots)
        total_issued = hd['total_issued']
        total_ccass = hd['total_ccass']
        ws.append([f"CCASS Holdings on {hd['date']}"])
        ws.append([])
        ws.append(['Type of holder', 'Holding', 'Stake%'])
        for label, val in [('Custodians', hd['custodians']), ('Brokers', hd['brokers']),
                           ('Other intermediaries', hd['other']), ('Intermediaries', hd['intermediaries']),
                           ('Total in CCASS', total_ccass),
                           ('Securities not in CCASS', total_issued - total_ccass),
                           ('Issued securities', total_issued)]:
            pct = f"{val / total_issued * 100:.2f}" if total_issued > 0 else "0.00"
            if label == 'Issued securities':
                pct = "100.00"
            ws.append([label, f"{val:,}", pct])
        ws.append([])
        ws.append(['Row', 'CCASS ID', 'Name', 'Holding', 'Lastchange', 'Stake%', 'Cumul.Stake%'])
        header_row_num = ws.max_row
        for cell in ws[header_row_num]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        cumul = 0.0
        for idx, p in enumerate(hd['participants'], 1):
            cumul += p['percentage']
            ws.append([idx, p['participant_id'], p.get('participant_name', ''),
                       f"{p['shareholding']:,}", hd['date'], f"{p['percentage']:.2f}", f"{cumul:.2f}"])
        for col_idx in range(1, 8):
            max_len = max(
                (len(str(c.value)) for r in ws.iter_rows(min_col=col_idx, max_col=col_idx) for c in r if c.value),
                default=0,
            )
            letter = chr(64 + col_idx)
            ws.column_dimensions[letter].width = max(40 if col_idx == 3 else 12, min(max_len + 2, 60))

        # ── Sheet 3 & 4: Buyer / Seller Movement ──
        for sheet_name, mvt_type, header_color, sub_color in [
            ("Buyer Movement", "buyer", "70AD47", "E2EFDA"),
            ("Seller Movement", "seller", "C00000", "FCE4D6"),
        ]:
            print(f"   Creating {sheet_name} sheet...")
            ws = wb.create_sheet(sheet_name)
            top_list, dates = self._create_movement_timeline(snapshots, top_n=20, movement_type=mvt_type)
            if not top_list:
                continue

            header1 = ['']
            header2 = ['Holdingdate']
            for _, data in top_list:
                header1.extend([data['name'], '', '|'])
                header2.extend(['Change', 'Stake%', '|'])
            ws.append(header1)
            ws.append(header2)
            for cell in ws[1]:
                if cell.value and cell.value != '|':
                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[2]:
                if cell.value:
                    if cell.value == '|':
                        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                    else:
                        cell.font = Font(bold=True, size=10)
                        cell.fill = PatternFill(start_color=sub_color, end_color=sub_color, fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            row_num = 3
            for date in reversed(dates):
                row = [date]
                for _, data in top_list:
                    change = data['changes'].get(date, 0)
                    stake = data['stakes'].get(date, 0)
                    row.extend([change, f"{stake:.2f}" if stake > 0 else '0', ''])
                ws.append(row)
                col_idx = 2
                for _, data in top_list:
                    change = data['changes'].get(date, 0)
                    ws.cell(row=row_num, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_num, column=col_idx + 1).alignment = Alignment(horizontal='center', vertical='center')
                    if change > 0:
                        ws.cell(row=row_num, column=col_idx).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        ws.cell(row=row_num, column=col_idx).font = Font(color='006100', bold=True)
                    elif change < 0:
                        ws.cell(row=row_num, column=col_idx).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        ws.cell(row=row_num, column=col_idx).font = Font(color='9C0006', bold=True)
                    ws.cell(row=row_num, column=col_idx + 2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                    col_idx += 3
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
                row_num += 1

            ws.column_dimensions['A'].width = 12
            for i in range(len(top_list)):
                ws.column_dimensions[get_column_letter(2 + i * 3)].width = 15
                ws.column_dimensions[get_column_letter(3 + i * 3)].width = 10
                ws.column_dimensions[get_column_letter(4 + i * 3)].width = 2

        # ── Individual participant sheets (B1-B10, S1-S10) ──
        print("   Creating individual participant sheets...")
        top_buyers, dates = self._create_movement_timeline(snapshots, top_n=20, movement_type='buyer')
        top_sellers, _ = self._create_movement_timeline(snapshots, top_n=20, movement_type='seller')

        for prefix, color, sub_color, plist in [
            ("B", "70AD47", "E2EFDA", top_buyers),
            ("S", "C00000", "FCE4D6", top_sellers),
        ]:
            for idx, (pid, data) in enumerate(plist[:10], 1):
                ws = wb.create_sheet(f"{prefix}{idx}")
                ws.append([f"{'Buyer' if prefix == 'B' else 'Seller'} {idx}: {data['name']}"])
                ws.merge_cells('A1:D1')
                ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
                ws['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 25
                ws.append([])
                ws.append(['Date', 'Change', 'Stake%', 'Holding'])
                for cell in ws[3]:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color=sub_color, end_color=sub_color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                r = 4
                for date in reversed(dates):
                    change = data['changes'].get(date, 0)
                    stake = data['stakes'].get(date, 0)
                    holding = data['holdings'].get(date, 0)
                    ws.append([date, f"{change:+,}" if change != 0 else '0', f"{stake:.2f}%", f"{holding:,}"])
                    if (prefix == 'B' and change > 0) or (prefix == 'S' and change < 0):
                        fill_c = 'C6EFCE' if prefix == 'B' else 'FFC7CE'
                        font_c = '006100' if prefix == 'B' else '9C0006'
                        ws.cell(row=r, column=2).fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type='solid')
                        ws.cell(row=r, column=2).font = Font(color=font_c, bold=True)
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    r += 1
                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 20

        wb.save(filepath)
        print(f"✅ Report saved: {filepath}")
        return filepath
