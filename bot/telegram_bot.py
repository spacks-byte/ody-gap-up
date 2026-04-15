#!/usr/bin/env python3
"""
Telegram bot that sends gap-up alerts and responds to commands.

Usage:
    1. Set environment variables:
         export TELEGRAM_BOT_TOKEN='your-token-from-BotFather'
         export TELEGRAM_CHAT_ID='your-chat-id'
    2. Run:  python bot.py

Commands the bot responds to:
    /scan   — run a gap-up scan now and send alerts
    /watch  — show watchlist summary
    /status — show bot status and last scan time
    /help   — list available commands
"""

import os
import logging
import threading
from datetime import datetime, timedelta, timezone
from pathlib import Path
from datetime import time as dt_time

# Load .env file if present (no extra dependency needed)
_env_path = Path(__file__).resolve().parent.parent / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip())

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ConversationHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

from scanner import TopGainersAnalyzer
from scoring import (
    save_prev_day_turnover,
    prefetch_yesterday_turnover,
    score_gap_up,
    score_intraday,
    get_prev_turnover,
    _format_turnover,
    _format_large_number,
    _get_yesterday_turnover,
    build_telegram_message,
    build_intraday_message,
)
from hkex.announcement_tracker import scan_announcements, scan_by_date, format_alerts
from hkex.llm_classifier import analyze_deal
from hkex.attention_list import add_from_scan_results, get_all as get_attention_stocks, get_codes as get_attention_codes, get_annotation
from bot.sheets_writer import write_scanner_hits, write_corporate_actions, write_ipo_tracker
from hkex.ipo_tracker import fetch_recent_ipos

# ── Config ──────────────────────────────────────────────────────────────────────

BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
FUTU_HOST = os.environ.get("FUTU_HOST", "127.0.0.1")
FUTU_PORT = int(os.environ.get("FUTU_PORT", "11111"))
SCAN_INTERVAL_S = int(os.environ.get("SCAN_INTERVAL", "300"))  # 5 min default
MARKET = os.environ.get("MARKET", "HK")
JOSLIST_FILE = Path(__file__).resolve().parent.parent / "joslist" / "fullstocks.txt"
JOSLIST_MANUAL_HEADING = "# Manual Addition"

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Shared state ────────────────────────────────────────────────────────────────

analyzer: TopGainersAnalyzer | None = None
last_scan_time: datetime | None = None
monitoring_active: bool = False
_eod_saved_date = None  # tracks whether we've saved today's EOD turnover


def get_analyzer() -> TopGainersAnalyzer:
    global analyzer
    if analyzer is None:
        analyzer = TopGainersAnalyzer(
            host=FUTU_HOST,
            port=FUTU_PORT,
            bot_token=BOT_TOKEN,
            chat_id=CHAT_ID,
        )
    return analyzer


# ── Market hours helper ─────────────────────────────────────────────────────────

def _hk_now():
    """Return (hm_int, datetime) in HKT. hm_int e.g. 930, 1545."""
    now_hkt = datetime.now(timezone.utc) + timedelta(hours=8)
    return now_hkt.hour * 100 + now_hkt.minute, now_hkt


def _hk_market_open() -> bool:
    hm, _ = _hk_now()
    return (930 <= hm < 1200) or (1300 <= hm < 1600)


def _normalize_joslist_code(raw: str) -> str | None:
    """Normalize user input into HK.XXXXX or US.SYMBOL format"""
    token = (raw or "").strip().upper()
    if not token:
        return None

    token = token.rstrip(",;")

    if token.endswith(".HK"):
        token = f"HK.{token[:-3]}"

    if token.startswith("HK."):
        suffix = token.split(".", 1)[1]
        digits = "".join(ch for ch in suffix if ch.isdigit())
        if not digits:
            return None
        return f"HK.{digits.zfill(5)}"

    if token.startswith("US."):
        symbol = token.split(".", 1)[1].strip()
        if not symbol:
            return None
        return f"US.{symbol}"

    if token.isdigit():
        return f"HK.{token.zfill(5)}"

    return None


def _load_joslist_lines() -> list[str]:
    return JOSLIST_FILE.read_text(encoding="utf-8").splitlines()


def _save_joslist_lines(lines: list[str]) -> None:
    JOSLIST_FILE.write_text("\n".join(lines).rstrip("\n") + "\n", encoding="utf-8")


def _manual_section_insert_index(lines: list[str]) -> int:
    heading_idx = next(
        (i for i, line in enumerate(lines) if line.strip().lower() == JOSLIST_MANUAL_HEADING.lower()),
        -1,
    )
    if heading_idx == -1:
        if lines and lines[-1].strip():
            lines.append("")
        lines.append(JOSLIST_MANUAL_HEADING)
        return len(lines)

    for i in range(heading_idx + 1, len(lines)):
        if lines[i].strip().startswith("#"):
            return i
    return len(lines)


# ── Telegram command handlers ───────────────────────────────────────────────────


def _sync_scanner_hits_to_sheets(analyzer):
    """Sync all tracked scanner stocks to Google Sheets 'Scanner Hits' tab."""
    import pandas as pd

    tracked = analyzer.load_tracked_stocks()
    market_stocks = {c: d for c, d in tracked.items()
                     if d.get('market') == MARKET}
    if not market_stocks:
        return None

    stock_codes = list(market_stocks.keys())
    all_snaps = []
    for i in range(0, len(stock_codes), 400):
        batch = stock_codes[i:i + 400]
        try:
            ret, snap = analyzer.quote_ctx.get_market_snapshot(batch)
            if ret == 0 and snap is not None and not snap.empty:
                all_snaps.append(snap)
        except Exception:
            pass

    snap_map = {}
    if all_snaps:
        snapshot = pd.concat(all_snaps, ignore_index=True)
        for _, row in snapshot.iterrows():
            snap_map[row['code']] = row

    rows = []
    for code, td in market_stocks.items():
        ini = td['initial_price']
        first_seen = datetime.fromisoformat(td['first_seen'])
        days = (datetime.now() - first_seen).days
        pinned = 'Yes' if td.get('permanent') else ''
        r = snap_map.get(code)
        cur = float(r['last_price']) if r is not None else ini
        daily_chg = float(r.get('change_rate', 0) or 0) if r is not None else 0.0
        volume = int(r.get('volume', 0) or 0) if r is not None else 0
        turnover = float(r.get('turnover', 0) or 0) if r is not None else 0.0
        total_chg = ((cur - ini) / ini * 100) if ini > 0 else 0

        rows.append({
            'Code': code,
            'Name': td['name'],
            'Initial Price': round(ini, 3),
            'Current Price': round(cur, 3),
            'Since Tracked %': round(total_chg, 1),
            'Daily Change %': round(daily_chg, 1),
            'Volume': volume,
            'Turnover (HK$)': round(turnover),
            'Days Tracked': days,
            'First Seen': first_seen.strftime('%Y-%m-%d %H:%M'),
            'Pinned': pinned,
        })

    rows.sort(key=lambda r: abs(r['Since Tracked %']), reverse=True)

    description = (
        "--- SCANNER HITS ---",
        "Stocks automatically flagged by the gap-up / intraday scanner (score >= 4).",
        "Auto-tracked for 3 days, then removed. Pinned stocks stay indefinitely.",
        f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')} HKT",
    )
    url = write_scanner_hits(rows, "\n".join(description))
    return url, rows


def _build_scan_excel(alerts, watches, tracked_rows, scan_type="intraday"):
    """
    Build an Excel workbook with up to 3 sheets:
      1) Alerts (score >= 8)
      2) Watch List (4 <= score < 8)
      3) Tracked Gainers (continued gainers from the tracked list)

    Args:
        alerts:  list of (row, result) tuples from the scanner
        watches: list of (row, result) tuples from the scanner
        tracked_rows: list of dicts with tracked gainer data
        scan_type: 'intraday' or 'gapup'

    Returns an io.BytesIO buffer with the .xlsx file.
    """
    import io
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Styling constants ──
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alert_fill = PatternFill("solid", fgColor="FCE4EC")
    alert_font = Font(color="B71C1C")
    watch_fill = PatternFill("solid", fgColor="FFF3E0")
    watch_font = Font(color="E65100")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006")
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    def _extract_rows(items, scan_type):
        """Convert (row, result) tuples into flat dicts for the DataFrame."""
        out = []
        for row, result in items:
            sig = result["signals"]
            code = result["code"]
            name = str(row.get("name", ""))[:25]
            price = float(row.get("last_price", 0) or 0)
            prev = float(row.get("prev_close_price", 0) or 0)
            daily_chg = ((price - prev) / prev * 100) if prev > 0 else 0
            open_p = float(row.get("open_price", 0) or 0)
            turnover = float(row.get("turnover", 0) or 0)
            mkt_cap = float(row.get("total_market_val", 0) or 0)

            d = {
                "Code": code,
                "Name": name,
                "Price": price,
                "Daily %": daily_chg,
                "Score": result["score"],
                "Tier": result["alert_tier"].upper(),
                "Cap Tier": sig.get("cap_tier", ""),
            }

            if scan_type == "gapup":
                d["Gap %"] = sig.get("gap_pct", 0)
                d["Continuation %"] = sig.get("continuation_pct", 0)
                d["Trigger"] = sig.get("trigger", "gap_up")
            else:
                d["Intraday %"] = sig.get("intraday_pct", 0)

            d["5min %"] = sig.get("change_5min", 0)
            d["Strength"] = sig.get("price_strength", 0)
            d["Delta %"] = sig.get("turnover_delta_pct", 0)
            d["Vol x Yday"] = sig.get("unusual_multiplier", 0)
            d["Turnover"] = turnover
            d["Market Cap"] = mkt_cap
            out.append(d)
        return out

    alert_rows = _extract_rows(alerts, scan_type)
    watch_rows = _extract_rows(watches, scan_type)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Alerts ──
        df_alerts = pd.DataFrame(alert_rows) if alert_rows else pd.DataFrame(columns=["No alerts found"])
        df_alerts.to_excel(writer, sheet_name="Alerts", index=False)

        # ── Sheet 2: Watch List ──
        df_watches = pd.DataFrame(watch_rows) if watch_rows else pd.DataFrame(columns=["No watch-tier stocks"])
        df_watches.to_excel(writer, sheet_name="Watch List", index=False)

        # ── Sheet 3: Tracked Gainers ──
        df_tracked = pd.DataFrame(tracked_rows) if tracked_rows else pd.DataFrame(columns=["No tracked gainers"])
        df_tracked.to_excel(writer, sheet_name="Tracked Gainers", index=False)

        # ── Styling helper ──
        def _style_sheet(ws, data_rows, row_fill=None, row_font=None):
            # Headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

            # Auto-width
            for col_idx in range(1, ws.max_column + 1):
                h = headers[col_idx - 1] or ""
                max_len = len(str(h)) + 4
                for row_idx in range(2, min(ws.max_row + 1, 50)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 22)

            if not data_rows:
                return

            # Format data cells
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    h = headers[col_idx - 1]

                    if h == "Price":
                        cell.number_format = "#,##0.000"
                        cell.alignment = Alignment(horizontal="right")
                    elif h in ("Daily %", "Gap %", "Continuation %",
                               "Intraday %", "5min %", "Delta %",
                               "Since Tracked %", "Today %"):
                        if cell.value is not None:
                            cell.value = cell.value / 100
                        cell.number_format = "+0.0%;-0.0%"
                        cell.alignment = Alignment(horizontal="center")
                    elif h == "Strength":
                        cell.number_format = "0%"
                        cell.alignment = Alignment(horizontal="center")
                    elif h == "Vol x Yday":
                        cell.number_format = "0.0"
                        cell.alignment = Alignment(horizontal="center")
                    elif h in ("Turnover", "Market Cap"):
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                    elif h == "Score":
                        cell.alignment = Alignment(horizontal="center")

                # Apply row tint
                if row_fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = row_fill
                        ws.cell(row=row_idx, column=col_idx).font = row_font

            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

        _style_sheet(writer.sheets["Alerts"], alert_rows, alert_fill, alert_font)
        _style_sheet(writer.sheets["Watch List"], watch_rows, watch_fill, watch_font)

        # Tracked Gainers: colour by today's change
        ws_t = writer.sheets["Tracked Gainers"]
        _style_sheet(ws_t, tracked_rows)  # base styling
        if tracked_rows:
            t_headers = [ws_t.cell(row=1, column=c).value
                         for c in range(1, ws_t.max_column + 1)]
            today_col = (t_headers.index("Today %") + 1) if "Today %" in t_headers else None
            if today_col:
                for row_idx in range(2, ws_t.max_row + 1):
                    val = ws_t.cell(row=row_idx, column=today_col).value
                    if val is not None:
                        tint_fill = green_fill if val > 0 else red_fill if val < 0 else None
                        tint_font = green_font if val > 0 else red_font if val < 0 else None
                        if tint_fill:
                            for c in range(1, ws_t.max_column + 1):
                                ws_t.cell(row=row_idx, column=c).fill = tint_fill
                                ws_t.cell(row=row_idx, column=c).font = tint_font

    buf.seek(0)
    return buf


def _get_tracked_gainers(analyzer, market):
    """
    Fetch live data for tracked stocks and return rows for continued gainers
    (stocks that are up today AND up since first tracked).
    """
    tracked = analyzer.load_tracked_stocks()
    market_stocks = {c: d for c, d in tracked.items()
                     if d.get("market") == market}
    if not market_stocks:
        return []

    codes = list(market_stocks.keys())
    all_snaps = []
    for i in range(0, len(codes), 400):
        batch = codes[i:i + 400]
        ret, snap = analyzer.quote_ctx.get_market_snapshot(batch)
        if ret == 0 and snap is not None and not snap.empty:
            all_snaps.append(snap)
    if not all_snaps:
        return []

    import pandas as pd
    snapshot = pd.concat(all_snaps, ignore_index=True)

    rows = []
    for _, row in snapshot.iterrows():
        code = row["code"]
        if code not in market_stocks:
            continue
        td = market_stocks[code]
        cur = float(row.get("last_price", 0) or 0)
        ini = td["initial_price"]
        if ini <= 0 or cur <= 0:
            continue

        prev_close = float(row.get("prev_close_price", 0) or 0)
        daily_chg = ((cur - prev_close) / prev_close * 100) if prev_close > 0 else 0
        total_chg = (cur - ini) / ini * 100
        turnover = float(row.get("turnover", 0) or 0)
        first_seen = datetime.fromisoformat(td["first_seen"])
        days = (datetime.now() - first_seen).days

        # Only include stocks that are UP today and have a positive total gain
        if daily_chg > 0 and total_chg > 0:
            rows.append({
                "Code": code,
                "Name": td["name"][:25],
                "Price": cur,
                "Today %": daily_chg,
                "Since Tracked %": total_chg,
                "Initial Price": ini,
                "Turnover": turnover,
                "Days Tracked": days,
            })

    rows.sort(key=lambda r: r["Since Tracked %"], reverse=True)
    return rows


async def cmd_scan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Run an intraday mover scan → Excel output with alerts + tracked gainers."""
    global last_scan_time
    hm, now_hkt = _hk_now()
    market_open = (930 <= hm < 1200) or (1300 <= hm < 1600)

    a = get_analyzer()

    if not market_open:
        time_str = now_hkt.strftime("%H:%M")
        await update.message.reply_text(
            f"🕐 HK market is closed ({time_str} HKT).\n"
            f"Generating end-of-market recap..."
        )
        try:
            msgs = a.get_eod_recap(market=MARKET, top_n=10)
            last_scan_time = datetime.now()
            if msgs:
                for msg in msgs:
                    await update.message.reply_text(msg, parse_mode="Markdown")
            else:
                await update.message.reply_text("Could not retrieve market data for recap.")
        except Exception as e:
            logger.exception("EOD recap failed")
            await update.message.reply_text(f"Recap error: {e}")
        return

    await update.message.reply_text(f"⏳ Scanning {MARKET} (intraday + gap-up)...")
    try:
        # Run both scanners, deduplicate (same logic as scheduled_scan)
        intra_alerts, intra_watches = a.scan_intraday_movers(market=MARKET)
        gap_alerts, gap_watches = a.scan_gap_ups(market=MARKET)

        seen = set()
        alerts, watches = [], []
        for row, result in intra_alerts:
            seen.add(result["code"])
            alerts.append((row, result))
        for row, result in gap_alerts:
            if result["code"] not in seen:
                seen.add(result["code"])
                alerts.append((row, result))
        for row, result in intra_watches:
            if result["code"] not in seen:
                seen.add(result["code"])
                watches.append((row, result))
        for row, result in gap_watches:
            if result["code"] not in seen:
                seen.add(result["code"])
                watches.append((row, result))

        # Track alerted & watch stocks
        for row, result in alerts + watches:
            code = result["code"]
            tracked = a.load_tracked_stocks()
            if code not in tracked:
                a.add_to_tracking(code, str(row.get("name", "")),
                                  float(row.get("last_price", 0) or 0),
                                  float(row.get("change_5min", 0) or 0), MARKET)

        last_scan_time = datetime.now()

        # Send top 2 alerts to Telegram
        top_alerts = sorted(alerts, key=lambda x: x[1].get("score", 0), reverse=True)[:2]
        for row, result in top_alerts:
            trigger = result.get("trigger", "gap_up")
            if trigger == "gap_up":
                msg = build_telegram_message(row, result)
            else:
                msg = build_intraday_message(row, result)
            await update.message.reply_text(msg, parse_mode="Markdown")

        # Sync all results to Google Sheets
        try:
            result_info = _sync_scanner_hits_to_sheets(a)
            if result_info:
                url, sheet_rows = result_info
                summary = (
                    f"📊 *Scan Complete*\n"
                    f"🔥 {len(alerts)} alerts  |  📊 {len(watches)} watch\n\n"
                    f"[Full results in Google Sheets]({url})"
                )
            else:
                summary = (
                    f"📊 *Scan Complete*\n"
                    f"🔥 {len(alerts)} alerts  |  📊 {len(watches)} watch"
                )
            await update.message.reply_text(summary, parse_mode="Markdown",
                                            disable_web_page_preview=True)
        except Exception as e:
            logger.warning("Scanner hits Sheets sync failed: %s", e)
            await update.message.reply_text(
                f"📊 *Scan Complete* — {len(alerts)} alerts, {len(watches)} watch\n"
                f"⚠️ Google Sheets sync failed.",
                parse_mode="Markdown",
            )
    except Exception as e:
        logger.exception("Scan failed")
        await update.message.reply_text(f"Scan error: {e}")



async def cmd_joslist(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Run Jo's tracker script with live progress updates."""
    import asyncio
    import subprocess
    import sys
    import re

    completion_marker = "JOSLIST_DONE"

    script_dir = Path(__file__).resolve().parent.parent / "joslist"
    script_path = script_dir / "stock_tracker.py"

    if not script_path.exists():
        await update.message.reply_text("❌ /joslist script not found at joslist/stock_tracker.py")
        return

    # Count total stocks for progress tracking
    stocks_file = script_dir / "fullstocks.txt"
    total_stocks = 0
    if stocks_file.exists():
        for line in stocks_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and (line.startswith("HK.") or line.startswith("US.")):
                total_stocks += 1

    progress_msg = await update.message.reply_text(
        f"⏳ Running /joslist — 0/{total_stocks} stocks processed..."
    )

    def _run_joslist_stream():
        """Run script and stream stdout lines back."""
        proc = subprocess.Popen(
            [sys.executable, "-u", str(script_path)],
            cwd=str(script_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        return proc

    loop = asyncio.get_event_loop()
    try:
        proc = await loop.run_in_executor(None, _run_joslist_stream)
    except Exception as e:
        logger.exception("/joslist script failed to start")
        await progress_msg.edit_text(f"❌ /joslist script failed to start: {e}")
        return

    processed = 0
    last_update_count = 0
    all_output = []
    found_marker = False
    update_interval = max(1, total_stocks // 20)  # update ~20 times

    def _read_line():
        return proc.stdout.readline()

    while True:
        line = await loop.run_in_executor(None, _read_line)
        if not line:
            break
        all_output.append(line)

        if completion_marker in line:
            found_marker = True

        # Count "Processing HK.xxxxx..." lines
        if line.strip().startswith("Processing "):
            processed += 1
            if processed - last_update_count >= update_interval:
                last_update_count = processed
                pct = int(processed / total_stocks * 100) if total_stocks else 0
                bar_fill = pct // 5
                bar = "█" * bar_fill + "░" * (20 - bar_fill)
                try:
                    await progress_msg.edit_text(
                        f"⏳ /joslist — {processed}/{total_stocks} stocks\n"
                        f"[{bar}] {pct}%"
                    )
                except Exception:
                    pass

    await loop.run_in_executor(None, proc.wait)

    if proc.returncode == 0 and found_marker:
        await progress_msg.edit_text(
            f"✅ /joslist finished — {processed}/{total_stocks} stocks processed.\n"
            "📄 Data is available here: "
            "https://docs.google.com/spreadsheets/d/1PiUuP3MNPPHUWVbLNQuPEWgABxa-8NNGiBpGG88EPEI/edit?gid=431873583#gid=431873583"
        )
        return

    stdout_text = "".join(all_output)
    if proc.returncode == 0 and not found_marker:
        output_tail = "\n".join(stdout_text.splitlines()[-20:])
        await progress_msg.edit_text(
            "❌ /joslist ended without completion marker, so success was not confirmed.\n"
            "Please check script output/logs and try again."
        )
        if output_tail:
            await update.message.reply_text(f"Last output:\n{output_tail}")
        return

    output_tail = "\n".join(stdout_text.splitlines()[-20:])
    if output_tail:
        await progress_msg.edit_text(
            f"❌ /joslist script failed (exit code {proc.returncode}).\n\n"
            f"Last output:\n{output_tail}"
        )
    else:
        await progress_msg.edit_text(
            f"❌ /joslist script failed (exit code {proc.returncode})."
        )


async def cmd_joadd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Add a stock code into joslist/fullstocks.txt under Manual Addition."""
    if not ctx.args:
        await update.message.reply_text("Usage: /joadd HK.00700  (or /joadd 700)")
        return

    code = _normalize_joslist_code(ctx.args[0])
    if not code:
        await update.message.reply_text("❌ Invalid code format. Use HK.00700, 700, or US.TSLA")
        return

    if not JOSLIST_FILE.exists():
        await update.message.reply_text("❌ joslist/fullstocks.txt was not found.")
        return

    try:
        lines = _load_joslist_lines()
        existing = {
            line.strip().upper()
            for line in lines
            if line.strip() and not line.strip().startswith("#")
        }

        if code in existing:
            await update.message.reply_text(f"ℹ️ {code} is already in the list.")
            return

        insert_at = _manual_section_insert_index(lines)
        lines.insert(insert_at, code)
        _save_joslist_lines(lines)
        await update.message.reply_text(f"✅ Added {code} to joslist (Manual Addition).")
    except Exception as e:
        logger.exception("/joadd failed")
        await update.message.reply_text(f"❌ /joadd failed: {e}")


async def cmd_joremove(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Remove a stock code from all occurrences in joslist/fullstocks.txt."""
    if not ctx.args:
        await update.message.reply_text("Usage: /joremove HK.00700  (or /joremove 700)")
        return

    code = _normalize_joslist_code(ctx.args[0])
    if not code:
        await update.message.reply_text("❌ Invalid code format. Use HK.00700, 700, or US.TSLA")
        return

    if not JOSLIST_FILE.exists():
        await update.message.reply_text("❌ joslist/fullstocks.txt was not found.")
        return

    try:
        lines = _load_joslist_lines()
        filtered = []
        removed = 0

        for line in lines:
            stripped = line.strip()
            if stripped and not stripped.startswith("#") and stripped.upper() == code:
                removed += 1
                continue
            filtered.append(line)

        if removed == 0:
            await update.message.reply_text(f"ℹ️ {code} was not found in joslist.")
            return

        _save_joslist_lines(filtered)
        await update.message.reply_text(f"✅ Removed {code} from joslist ({removed} entr{'y' if removed == 1 else 'ies'}).")
    except Exception as e:
        logger.exception("/joremove failed")
        await update.message.reply_text(f"❌ /joremove failed: {e}")


# ── Debug conversation states ───────────────────────────────────────────────────
DEBUG_CODE = 0


def _debug_fetch_yesterday_turnover(futu_code: str) -> float:
    """
    Fetch yesterday's turnover from yfinance for a single stock.
    Returns volume × close for the previous trading day, or 0 on failure.
    Uses Ticker.history() instead of yf.download() for simpler column handling.
    """
    import logging as _logging
    _logging.getLogger("yfinance").setLevel(_logging.CRITICAL)

    try:
        import yfinance as yf
        # Convert HK.00022 → 0022.HK
        parts = futu_code.split(".")
        if len(parts) == 2 and parts[0] == "HK":
            num = parts[1].lstrip("0").zfill(4)
            yf_ticker = f"{num}.HK"
        else:
            return 0.0

        ticker = yf.Ticker(yf_ticker)
        df = ticker.history(period="5d")
        if df is None or df.empty or len(df) < 2:
            logger.warning("yfinance returned <2 rows for %s", yf_ticker)
            return 0.0

        row = df.iloc[-2]  # yesterday
        close_val = float(row["Close"])
        vol_val = float(row["Volume"])
        turnover = close_val * vol_val
        return turnover if turnover > 0 else 0.0
    except Exception as e:
        logger.warning("yfinance fetch failed for %s: %s", futu_code, e)
        return 0.0


async def cmd_debug(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Start debug flow — ask for stock code."""
    await update.message.reply_text(
        "🔍 *Debug Score Breakdown*\n\n"
        "Enter a stock code (e.g. `HK.01705`, `00700`, `1705`):",
        parse_mode="Markdown",
    )
    return DEBUG_CODE


async def debug_got_code(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Received stock code — fetch data and send fully detailed breakdown."""
    raw = update.message.text.strip()

    # Normalise to HK.XXXXX format
    raw = raw.replace('.HK', '').replace('.hk', '')
    if raw.upper().startswith('HK.'):
        code = raw.upper()
    else:
        digits = raw.lstrip('0') or '0'
        code = f"HK.{digits.zfill(5)}"

    await update.message.reply_text(f"⏳ Fetching snapshot for `{code}`…", parse_mode="Markdown")

    try:
        a = get_analyzer()
        ret, snap = a.quote_ctx.get_market_snapshot([code])
        if ret != 0 or snap is None or snap.empty:
            await update.message.reply_text(f"❌ Could not fetch data for `{code}`. Check the code and try again.", parse_mode="Markdown")
            return ConversationHandler.END

        row = snap.iloc[0]

        # ── Raw market data ──
        last_price    = float(row.get("last_price", 0) or 0)
        open_price    = float(row.get("open_price", 0) or 0)
        prev_close    = float(row.get("prev_close_price", 0) or 0)
        high_price    = float(row.get("high_price", 0) or 0)
        low_price     = float(row.get("low_price", 0) or 0)
        volume        = int(row.get("volume", 0) or 0)
        turnover      = float(row.get("turnover", 0) or 0)
        market_cap    = float(row.get("total_market_val", 0) or 0)
        change_rate   = float(row.get("change_rate", 0) or 0)
        change_5min   = float(row.get("change_5min", 0) or 0) if "change_5min" in row.index else 0.0
        name          = str(row.get("name", ""))[:25]

        # Previous scan turnover
        prev_scan_turnover = get_prev_turnover(MARKET, code)

        # ── Fetch real yesterday turnover from yfinance (EOD-safe) ──
        ytd_turnover_yf = _debug_fetch_yesterday_turnover(code)
        json_turnover = _get_yesterday_turnover(MARKET, code)
        yesterday_turnover = ytd_turnover_yf if ytd_turnover_yf > 0 else json_turnover
        ytd_source = "yfinance" if ytd_turnover_yf > 0 else "JSON (no yfinance data)"

        # ── Message 1: Raw input data ──
        data_msg = (
            f"📋 *Debug: {code}* — {name}\n"
            f"\n"
            f"*━━━ RAW MARKET DATA ━━━*\n"
            f"Last Price: `HK${last_price:.4f}`\n"
            f"Open Price: `HK${open_price:.4f}`\n"
            f"Prev Close: `HK${prev_close:.4f}`\n"
            f"High: `HK${high_price:.4f}`\n"
            f"Low:  `HK${low_price:.4f}`\n"
            f"Daily Change: `{change_rate:+.2f}%`\n"
            f"5-min Change: `{change_5min:+.2f}%`\n"
            f"Volume: `{volume:,}`\n"
            f"Turnover (today): `{_format_turnover(turnover)}`\n"
            f"Market Cap: `{_format_large_number(market_cap, 'HK$')}`\n"
            f"\n"
            f"*━━━ BASELINES ━━━*\n"
            f"Prev Scan Turnover: `{_format_turnover(prev_scan_turnover)}`\n"
            f"Yesterday Turnover: `{_format_turnover(yesterday_turnover)}` ({ytd_source})\n"
            f"JSON file value: `{_format_turnover(json_turnover)}`\n"
        )
        await update.message.reply_text(data_msg, parse_mode="Markdown")

        # ── Run INTRADAY scorer ──
        result_intra = score_intraday(row, market=MARKET,
                                      override_yesterday_turnover=yesterday_turnover)
        s = result_intra["signals"]

        # ── Message 2: Signal 1 — Intraday Move ──
        mt = s["move_thresholds"]  # tuple (t1, t2, t3)
        sig1_msg = (
            f"*━━━ INTRADAY SCORER ━━━*\n\n"
            f"*1️⃣ Intraday Move (0–3 pts)* → *{s['move_score']} pts*\n\n"
            f"Formula:\n"
            f"  `(last - open) / open × 100`\n"
            f"  `= ({last_price:.4f} - {open_price:.4f}) / {open_price:.4f} × 100`\n"
            f"  `= {s['intraday_pct']:+.2f}%`\n\n"
            f"Cap tier: *{s['cap_tier']}* (mkt cap = `{_format_large_number(market_cap, 'HK$')}`)\n"
            f"  large ≥ HK$5B  |  mid ≥ HK$500M  |  small < HK$500M\n\n"
            f"Thresholds for {s['cap_tier']}-cap:\n"
            f"  `≥ {mt[0]}%` → 1pt  |  `≥ {mt[1]}%` → 2pts  |  `≥ {mt[2]}%` → 3pts\n"
            f"  `{s['intraday_pct']:+.2f}%` vs `{mt[0]}%/{mt[1]}%/{mt[2]}%`  →  *{s['move_score']} pts*\n\n"
            f"*2️⃣ Daily Change (-1 to 2 pts)* → *{s['daily_score']} pts*\n\n"
            f"Formula:\n"
            f"  `(last - prev_close) / prev_close × 100`\n"
            f"  `= ({last_price:.4f} - {prev_close:.4f}) / {prev_close:.4f} × 100`\n"
            f"  `= {s['daily_chg']:+.2f}%`\n\n"
        )
        dt = s["daily_thresholds"]  # tuple (neg, t1, t2)
        sig1_msg += (
            f"Thresholds for {s['cap_tier']}-cap:\n"
            f"  `< {dt[0]}%` → -1pt  |  `≥ {dt[1]}%` → +1pt  |  `≥ {dt[2]}%` → +2pts\n"
            f"  `{s['daily_chg']:+.2f}%` vs thresholds  →  *{s['daily_score']} pts*"
        )
        await update.message.reply_text(sig1_msg, parse_mode="Markdown")

        # ── Message 3: Signal 3+4 — Momentum + Strength ──
        range_span = high_price - low_price
        sig34_msg = (
            f"*3️⃣ 5-min Momentum (-1 to 1 pt)* → *{s['momentum_score']} pts*\n\n"
            f"Raw value from Futu: `{s['change_5min']:+.2f}%`\n"
            f"  `< -1%` → -1pt  |  `≥ 3%` → +1pt  |  else → 0\n"
            f"  `{s['change_5min']:+.2f}%`  →  *{s['momentum_score']} pts*\n\n"
            f"*4️⃣ Price Strength (-1 to 1 pt)* → *{s['strength_score']} pts*\n\n"
            f"Formula:\n"
            f"  `(last - low) / (high - low)`\n"
            f"  `= ({last_price:.4f} - {low_price:.4f}) / ({high_price:.4f} - {low_price:.4f})`\n"
        )
        if range_span > 0:
            sig34_msg += f"  `= {last_price - low_price:.4f} / {range_span:.4f}`\n"
        sig34_msg += (
            f"  `= {s['price_strength']:.2f}`\n\n"
            f"  `< 0.25` → -1pt  |  `≥ 0.75` → +1pt  |  else → 0\n"
            f"  `{s['price_strength']:.2f}`  →  *{s['strength_score']} pts*"
        )
        await update.message.reply_text(sig34_msg, parse_mode="Markdown")

        # ── Message 4: Signal 5 — Turnover Delta ──
        delta_src = "turnover_now" if s["prev_scan_turnover"] == 0 else "turnover_now - prev_scan"
        sig5_msg = (
            f"*5️⃣ Turnover Delta (0–3 pts)* → *{s['delta_score']} pts*\n\n"
            f"Step 1 — New flow since last scan:\n"
        )
        if s["prev_scan_turnover"] == 0:
            sig5_msg += (
                f"  No prev scan → delta = today's full turnover\n"
                f"  `delta = {_format_turnover(s['turnover_now'])}`\n\n"
            )
        else:
            sig5_msg += (
                f"  `max(0, turnover_now - prev_scan_turnover)`\n"
                f"  `= max(0, {_format_turnover(s['turnover_now'])} - {_format_turnover(s['prev_scan_turnover'])})`\n"
                f"  `= {_format_turnover(s['turnover_delta_hkd'])}`\n\n"
            )

        sig5_msg += (
            f"Step 2 — Elapsed session fraction:\n"
            f"  `min(elapsed_min, 330) / 330`\n"
            f"  `= min({s['elapsed_min']:.1f}, 330) / 330`\n"
            f"  `= {s['elapsed_fraction']:.4f}`\n\n"
            f"Step 3 — Expected turnover (baseline):\n"
        )
        if s["yesterday_turnover"] > 0:
            sig5_msg += f"  Using yesterday's turnover: `{_format_turnover(s['yesterday_turnover'])}`\n\n"
        else:
            sig5_msg += (
                f"  No yesterday data → using projected today\n"
                f"  `turnover_now / elapsed_frac = {_format_turnover(s['turnover_now'])} / {s['elapsed_fraction']:.4f}`\n"
                f"  `= {_format_turnover(s['expected_turnover'])}`\n\n"
            )

        sig5_msg += (
            f"Step 4 — Delta %:\n"
            f"  `delta / expected × 100`\n"
            f"  `= {_format_turnover(s['turnover_delta_hkd'])} / {_format_turnover(s['expected_turnover'])} × 100`\n"
            f"  `= {s['turnover_delta_pct']:.1f}%`\n\n"
            f"Cap tier: *{s['cap_tier']}*\n"
            f"Thresholds: `≥ {s['delta_thresholds'][0]}%` → 1pt  |  `≥ {s['delta_thresholds'][1]}%` → 2pts  |  `≥ {s['delta_thresholds'][2]}%` → 3pts\n"
            f"  `{s['turnover_delta_pct']:.1f}%`  →  *{s['delta_score']} pts*"
        )
        await update.message.reply_text(sig5_msg, parse_mode="Markdown")

        # ── Message 6: Final result + veto check ──
        tier_emoji = {"alert": "🔥", "watch": "📊", "skip": "⬜"}.get(result_intra["alert_tier"], "⬜")
        result_msg = (
            f"*━━━ FINAL SCORE ━━━*\n\n"
            f"Raw sum:\n"
            f"  `{s['move_score']} + {s['daily_score']} + {s['momentum_score']}"
            f" + {s['strength_score']} + {s['delta_score']} + {s['unusual_score']}`\n"
            f"  `= {s['raw_score']}`\n\n"
        )
        if s["veto_applied"]:
            if s["veto_applied"] == "ghost_trade":
                result_msg += (
                    f"⚠️ *VETO: Ghost Trade*\n"
                    f"  delta% `{s['turnover_delta_pct']:.1f}%` < `{s['delta_veto']}%` (veto threshold for {s['cap_tier']}-cap)\n"
                    f"  AND turnover_now/yesterday < 0.5\n"
                    f"  → Score capped at 3\n\n"
                )
            elif s["veto_applied"] == "illiquid":
                result_msg += (
                    f"⚠️ *VETO: Illiquid*\n"
                    f"  turnover `{_format_turnover(s['turnover_now'])}` < HK$50k\n"
                    f"  → Score capped at 3\n\n"
                )
        else:
            result_msg += f"No veto applied.\n\n"

        result_msg += (
            f"Final: `clamp(0, 12, {s['raw_score']})` = *{result_intra['score']}*\n\n"
            f"{tier_emoji} Score: *{result_intra['score']}/12*\n"
            f"Tier: *{result_intra['alert_tier'].upper()}*\n"
            f"  (alert ≥ 8 | watch ≥ 4 | skip < 4)"
        )
        await update.message.reply_text(result_msg, parse_mode="Markdown")

        # ── B) Run GAP-UP scorer if stock actually gapped ──
        is_gap_up = (open_price > prev_close > 0)
        if is_gap_up:
            rg = score_gap_up(row, market=MARKET, trigger="gap_up",
                              override_yesterday_turnover=yesterday_turnover)
            g = rg["signals"]
            gt = g["gap_thresholds"]  # tuple

            gap_msg1 = (
                f"*━━━ GAP-UP SCORER ━━━*\n\n"
                f"*1️⃣ Overnight Gap (0–3 pts)* → *{g['gap_score']} pts*\n\n"
                f"Formula:\n"
                f"  `(open - prev_close) / prev_close × 100`\n"
                f"  `= ({open_price:.4f} - {prev_close:.4f}) / {prev_close:.4f} × 100`\n"
                f"  `= {g['gap_pct']:+.2f}%`\n\n"
                f"Cap tier: *{g['cap_tier']}*\n"
                f"Thresholds for {g['cap_tier']}-cap:\n"
                f"  `≥ {gt[0]}%` → 1pt  |  `≥ {gt[1]}%` → 2pts  |  `≥ {gt[2]}%` → 3pts\n"
                f"  `{g['effective_gap_pct']:+.2f}%` vs thresholds  →  *{g['gap_score']} pts*\n\n"
                f"*2️⃣ Continuation (-1 to 2 pts)* → *{g['cont_score']} pts*\n\n"
                f"Formula:\n"
                f"  `(last - open) / open × 100`\n"
                f"  `= ({last_price:.4f} - {open_price:.4f}) / {open_price:.4f} × 100`\n"
                f"  `= {g['continuation_pct']:+.2f}%`\n\n"
                f"  `< -3%` → -1pt  |  `0–3%` → +1pt  |  `≥ 3%` → +2pts\n"
                f"  `{g['continuation_pct']:+.2f}%`  →  *{g['cont_score']} pts*\n\n"
                f"Signals 3–6 same as intraday.\n\n"
            )

            # Final gap score
            gtier_emoji = {"alert": "🔥", "watch": "📊", "skip": "⬜"}.get(rg["alert_tier"], "⬜")
            gap_msg1 += (
                f"*━━━ GAP-UP RESULT ━━━*\n\n"
                f"Raw: `{g['gap_score']} + {g['cont_score']} + {g['momentum_score']}"
                f" + {g['strength_score']} + {g['delta_score']} + {g['unusual_score']}`"
            )
            if g["breakout_bonus"]:
                gap_msg1 += f" + `{g['breakout_bonus']}` (breakout bonus)"
            gap_msg1 += f"\n  `= {g['raw_score']}`\n"
            if g["veto_applied"]:
                gap_msg1 += f"⚠️ Veto: {g['veto_applied']} → capped\n"
            gap_msg1 += (
                f"\n{gtier_emoji} Gap-Up Score: *{rg['score']}/12*  "
                f"Tier: *{rg['alert_tier'].upper()}*"
            )
            await update.message.reply_text(gap_msg1, parse_mode="Markdown")

    except Exception as e:
        logger.exception("Debug failed")
        await update.message.reply_text(f"❌ Error: {e}")

    return ConversationHandler.END


async def debug_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Debug cancelled.")
    return ConversationHandler.END


# ── CCASS conversation states ───────────────────────────────────────────────────
CCASS_CODE, CCASS_SOURCE, CCASS_PERIOD, CCASS_INTERVAL = range(4)


async def cmd_ccass(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Start CCASS report flow — ask for stock code."""
    await update.message.reply_text(
        "📊 *CCASS Shareholding Report*\n\n"
        "Enter the stock code (e.g. `00700`, `1985`):",
        parse_mode="Markdown",
    )
    return CCASS_CODE


async def ccass_got_code(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Received stock code — ask for data source."""
    raw = update.message.text.strip().replace('.HK', '').replace('.hk', '')
    if not raw.isdigit():
        await update.message.reply_text("❌ Invalid code. Send a numeric stock code (e.g. `00700`):")
        return CCASS_CODE

    ctx.user_data['ccass_code'] = raw.zfill(5)
    await update.message.reply_text(
        "Choose data source:\n\n"
        "1️⃣  *HKEX* — official, last 12 months (recommended)\n"
        "2️⃣  *Webb-site* — historical, beyond 12 months\n"
        "⚠️ _Webb-site may be blocked (HTTP 403). Use HKEX if unsure._\n\n"
        "Reply *1* or *2*:",
        parse_mode="Markdown",
    )
    return CCASS_SOURCE


async def ccass_got_source(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Received source — ask for period."""
    choice = update.message.text.strip()
    if choice not in ('1', '2'):
        await update.message.reply_text("Reply *1* (HKEX) or *2* (Webb-site):", parse_mode="Markdown")
        return CCASS_SOURCE

    ctx.user_data['ccass_source'] = 'hkex' if choice == '1' else 'webb'
    await update.message.reply_text(
        "Choose period:\n\n"
        "1️⃣  1 month\n"
        "2️⃣  3 months\n"
        "3️⃣  6 months\n"
        "4️⃣  12 months\n\n"
        "Reply *1–4*:",
        parse_mode="Markdown",
    )
    return CCASS_PERIOD


async def ccass_got_period(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Received period — ask for interval."""
    choice = update.message.text.strip()
    period_map = {'1': 30, '2': 90, '3': 180, '4': 365}
    if choice not in period_map:
        await update.message.reply_text("Reply *1–4*:", parse_mode="Markdown")
        return CCASS_PERIOD

    ctx.user_data['ccass_days'] = period_map[choice]
    await update.message.reply_text(
        "Choose snapshot interval:\n\n"
        "1️⃣  Daily\n"
        "2️⃣  Weekly (7 days)\n"
        "3️⃣  Bi-weekly (14 days)\n"
        "4️⃣  Monthly (30 days)\n\n"
        "Reply *1–4*:",
        parse_mode="Markdown",
    )
    return CCASS_INTERVAL


async def ccass_got_interval(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Received interval — run analysis and send Excel."""
    choice = update.message.text.strip()
    interval_map = {'1': 1, '2': 7, '3': 14, '4': 30}
    if choice not in interval_map:
        await update.message.reply_text("Reply *1–4*:", parse_mode="Markdown")
        return CCASS_INTERVAL

    interval_days = interval_map[choice]
    stock_code = ctx.user_data['ccass_code']
    source = ctx.user_data['ccass_source']
    lookback = ctx.user_data['ccass_days']

    from datetime import datetime as _dt, timedelta as _td
    end_date = _dt.now().strftime("%Y/%m/%d")
    start_date = (_dt.now() - _td(days=lookback)).strftime("%Y/%m/%d")

    period_label = {30: '1 month', 90: '3 months', 180: '6 months', 365: '12 months'}[lookback]
    interval_label = {1: 'daily', 7: 'weekly', 14: 'bi-weekly', 30: 'monthly'}[interval_days]
    source_label = 'HKEX' if source == 'hkex' else 'Webb-site'

    await update.message.reply_text(
        f"⏳ Starting CCASS analysis…\n\n"
        f"Stock: `{stock_code}`\n"
        f"Source: {source_label}\n"
        f"Period: {period_label} ({start_date} → {end_date})\n"
        f"Interval: {interval_label}\n\n"
        f"This may take a few minutes. I'll send the report when ready.",
        parse_mode="Markdown",
    )

    # Run the heavy work in a thread so the bot stays responsive
    import asyncio
    chat_id = update.effective_chat.id
    bot = ctx.bot

    def _run_analysis():
        import io
        try:
            from ccass import HistoricalCCASSTracker
            tracker = HistoricalCCASSTracker(source=source)

            snapshot_dates = tracker.generate_snapshot_dates(
                start_date, end_date, interval_days
            )
            snapshots = tracker.fetch_historical_snapshots(
                stock_code, snapshot_dates, delay_seconds=1.5
            )

            if not snapshots:
                return None, "No data retrieved — check the stock code or try a different period."

            # Generate Excel to a temp dir
            import tempfile
            tmpdir = tempfile.mkdtemp(prefix="ccass_")
            report_path = tracker.generate_excel_report(
                stock_code, snapshots, start_date, output_dir=tmpdir
            )

            # Read the file into memory
            with open(report_path, 'rb') as f:
                buf = io.BytesIO(f.read())
            buf.seek(0)

            fname = os.path.basename(report_path)
            n_snaps = len(snapshots)
            return (buf, fname, n_snaps), None

        except Exception as e:
            logger.exception("CCASS analysis failed")
            return None, str(e)

    loop = asyncio.get_event_loop()
    result, error = await loop.run_in_executor(None, _run_analysis)

    if error:
        await bot.send_message(chat_id=chat_id, text=f"❌ CCASS error: {error}")
    else:
        buf, fname, n_snaps = result
        await bot.send_message(
            chat_id=chat_id,
            text=f"✅ CCASS report ready — {n_snaps} snapshots collected.",
        )
        await bot.send_document(
            chat_id=chat_id,
            document=buf,
            filename=fname,
            caption=f"CCASS Analysis: {stock_code} ({source_label}, {period_label})",
        )

    return ConversationHandler.END


async def ccass_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Cancel the CCASS conversation."""
    await update.message.reply_text("CCASS report cancelled.")
    return ConversationHandler.END


async def cmd_ipo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Fetch recent HK IPOs and write performance to Google Sheets."""
    await update.message.reply_text("Fetching recent IPOs from Futu...")

    try:
        a = get_analyzer()
        rows = fetch_recent_ipos(a.quote_ctx, market=MARKET)

        if not rows:
            await update.message.reply_text("No recent IPOs found in the last 90 days.")
            return

        description = (
            "--- IPO TRACKER ---",
            "Recent HK IPOs tracked for 90 days after listing.",
            "Data: Futu 'Recent IPOs' plate + market snapshot + industry plates.",
            "IPO Price = official offer price if available, otherwise first-day opening price.",
            f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')} HKT",
        )
        url = write_ipo_tracker(rows, "\n".join(description))

        # Count stats
        with_return = [r for r in rows if isinstance(r.get("Return Since IPO %"), (int, float))]
        n_up = sum(1 for r in with_return if r["Return Since IPO %"] > 0)
        n_down = sum(1 for r in with_return if r["Return Since IPO %"] < 0)

        await update.message.reply_text(
            f"*IPO Tracker* — {len(rows)} recent IPOs\n"
            f"📈 {n_up} up  📉 {n_down} down since listing\n\n"
            f"[Open in Google Sheets]({url})",
            parse_mode="Markdown",
            disable_web_page_preview=True,
        )

    except Exception as e:
        logger.exception("IPO tracker failed")
        await update.message.reply_text(f"Error: {e}")


async def cmd_monitor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Toggle market monitoring mode on/off."""
    global monitoring_active
    monitoring_active = not monitoring_active
    if monitoring_active:
        await update.message.reply_text(
            "🟢 *Market Monitoring ON*\n\n"
            f"Scanning {MARKET} every {SCAN_INTERVAL_S // 60} min during trading hours.\n"
            "I'll send alerts when I spot opportunities.",
            parse_mode="Markdown",
        )
    else:
        await update.message.reply_text(
            "🔴 *Market Monitoring OFF*\n\n"
            "Automatic scanning paused. Use /scan for manual checks.",
            parse_mode="Markdown",
        )


async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Report bot status."""
    tracked = get_analyzer().load_tracked_stocks()
    monitor_str = "🟢 ON" if monitoring_active else "🔴 OFF"
    lines = [
        f"*Bot Status*",
        f"Market: {MARKET}",
        f"Monitoring: {monitor_str}",
        f"Scan interval: {SCAN_INTERVAL_S // 60} min",
        f"Scanner hits: {len(tracked)}",
        f"Last scan: {last_scan_time.strftime('%H:%M:%S') if last_scan_time else 'never'}",
    ]
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_news(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Check HKEX for material announcements. Shows date picker or accepts date arg."""
    date_arg = " ".join(ctx.args).strip() if ctx.args else ""

    if date_arg:
        # Direct date provided — scan that date
        await _news_scan_date(update, date_arg)
    else:
        # No date — show buttons
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Today (last 24h)", callback_data="news|today")],
            [InlineKeyboardButton("Custom date", callback_data="news|custom")],
        ])
        await update.message.reply_text(
            "*HKEX Announcements*\nWhich date do you want to scan?",
            reply_markup=keyboard,
            parse_mode="Markdown",
        )


# State for /news custom date flow
NEWS_DATE = 990


async def news_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle /news inline button press."""
    query = update.callback_query
    await query.answer()
    data = query.data or ""

    if data == "news|today":
        await query.edit_message_text("🔎 Scanning HKEX announcements…")
        try:
            results = scan_announcements(since_hours=24.0)
            if results:
                text = format_alerts(results)
                for i in range(0, len(text), 4000):
                    chunk = text[i:i + 4000]
                    await query.message.reply_text(
                        chunk, parse_mode="Markdown",
                        disable_web_page_preview=True,
                    )
                added = add_from_scan_results(results)
                if added:
                    await query.message.reply_text(
                        f"📌 Added {len(added)} stock(s) to Corporate Actions watchlist.",
                    )
                await _auto_deal_analysis_from_msg(query.message, results)
            else:
                await query.message.reply_text(
                    "✅ No material announcements in the last 24 hours."
                )
        except Exception as e:
            logger.exception("News scan failed")
            await query.message.reply_text(f"❌ Scan failed: {e}")

    elif data == "news|custom":
        await query.edit_message_text(
            "📅 Send me a date in one of these formats:\n"
            "`DD/MM/YYYY`  or  `YYYY-MM-DD`\n\n"
            "_Note: only the last ~7 days are available._",
            parse_mode="Markdown",
        )
        ctx.user_data["awaiting_news_date"] = True


async def news_date_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle custom date text after user picks 'Custom date'."""
    if not ctx.user_data.get("awaiting_news_date"):
        return
    ctx.user_data["awaiting_news_date"] = False
    date_arg = update.message.text.strip()
    await _news_scan_date(update, date_arg)


async def _news_scan_date(update, date_arg: str):
    """Scan HKEX announcements for a specific date."""
    msg = update.message if hasattr(update, 'message') and update.message else update
    await msg.reply_text(f"🔎 Scanning HKEX announcements for *{date_arg}*…", parse_mode="Markdown")
    try:
        results = scan_by_date(date_arg)
        if results:
            text = format_alerts(results, date_label=date_arg)
            for i in range(0, len(text), 4000):
                chunk = text[i:i + 4000]
                await msg.reply_text(
                    chunk, parse_mode="Markdown",
                    disable_web_page_preview=True,
                )
            added = add_from_scan_results(results)
            if added:
                await msg.reply_text(
                    f"📌 Added {len(added)} stock(s) to Corporate Actions watchlist.",
                )
            await _auto_deal_analysis_from_msg(msg, results)
        else:
            await msg.reply_text(
                f"✅ No material announcements found for {date_arg}.\n"
                "_Note: only the last ~7 days are available._",
                parse_mode="Markdown",
            )
    except ValueError as e:
        await msg.reply_text(f"❌ {e}")
    except Exception as e:
        logger.exception("News scan failed")
        await msg.reply_text(f"❌ Scan failed: {e}")


async def _auto_deal_analysis_from_msg(message, results: list[dict]):
    """Auto-run deal analysis for Privatisation/Takeover/M&A items."""
    deal_items = [
        r for r in results
        if r.get("is_original", True)
        and r.get("category") in ("Privatisation", "Takeover", "M&A")
    ]
    if not deal_items:
        return

    for item in deal_items:
        code = item.get("stock_code", "")
        name = item.get("stock_name", "")
        cat = item.get("category", "")
        try:
            await message.reply_text(
                f"🔎 Auto-analysing deal: `{code}` *{name}* ({cat})…",
                parse_mode="Markdown",
            )
            analysis = analyze_deal(
                title=item.get("title", ""),
                stock_code=code,
                stock_name=name,
                link=item.get("link", ""),
            )
            if analysis:
                for i in range(0, len(analysis), 4000):
                    await message.reply_text(
                        analysis[i:i + 4000],
                        parse_mode="Markdown",
                        disable_web_page_preview=True,
                    )
            else:
                await message.reply_text(f"❌ Deal analysis failed for {code} {name}.")
        except Exception as e:
            logger.warning("Auto deal analysis failed for %s: %s", code, e)


# Module-level cache for pending deal analysis requests
_pending_deals: dict[str, dict] = {}


async def deal_analysis_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle inline button press for deal analysis."""
    query = update.callback_query
    await query.answer()

    data = query.data or ""
    if not data.startswith("deal|"):
        return

    news_id = data.split("|", 1)[1]
    item = _pending_deals.get(news_id)

    if not item:
        await query.edit_message_text("❌ Deal info expired. Run /news again.")
        return

    await query.edit_message_text(
        f"🔎 Analysing deal for `{item['stock_code']}` {item['stock_name']}…\n"
        "_Downloading PDF and running AI analysis…_",
        parse_mode="Markdown",
    )

    analysis = analyze_deal(
        title=item.get("title", ""),
        stock_code=item.get("stock_code", ""),
        stock_name=item.get("stock_name", ""),
        link=item.get("link", ""),
    )

    if analysis:
        # Split long messages (Telegram 4096 char limit)
        for i in range(0, len(analysis), 4000):
            chunk = analysis[i:i + 4000]
            if i == 0:
                await query.edit_message_text(
                    chunk, parse_mode="Markdown",
                    disable_web_page_preview=True,
                )
            else:
                await query.message.reply_text(
                    chunk, parse_mode="Markdown",
                    disable_web_page_preview=True,
                )
    else:
        await query.edit_message_text("❌ Deal analysis failed.")


async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    monitor_str = "ON" if monitoring_active else "OFF"
    text = (
        "*Commands*\n"
        "\n"
        "*Scan & Monitor*\n"
        "/scan — Market scan (intraday + gap-up)\n"
        "/monitor — Toggle auto-scan (" + monitor_str + ")\n"
        "\n"
        "*Lists (Google Sheets)*\n"
        "/corpactions — HKEX corporate actions (90-day)\n"
        "/ipo — Recent IPOs performance (90-day)\n"
        "\n"
        "*News & Research*\n"
        "/news — HKEX announcements (today or custom date)\n"
        "/debug — Score breakdown for any stock\n"
        "/ccass — CCASS shareholding report\n"
        "/chat — AI stock analyst\n"
        "\n"
        "*Jo's Tracker*\n"
        "/joslist — Run Jo's tracker script\n"
        "/joadd CODE — Add stock to Jo's list\n"
        "/joremove CODE — Remove from Jo's list\n"
        "\n"
        "*Info*\n"
        "/status — Bot status\n"
        "/algo — How scoring works\n"
        "/help — This message"
    )
    await update.message.reply_text(text, parse_mode="Markdown")


async def cmd_algo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Explain the scoring algorithm and reporting metrics."""
    part1 = (
        "*📐 Scoring Algorithm*\n"
        "\n"
        "Two scanning paths: *Gap-Up* (overnight) and "
        "*Intraday Breakout* (flat open → mid-session spike).\n"
        "All data from Futu snapshot — zero K-line quota.\n"
        "\n"
        "*━━━ PATH A: GAP-UP ━━━*\n"
        "Triggers when: `open > prev_close`\n"
        "Scored 0–12 across 6 signals. Alert at 8+, watch at 4+.\n"
        "\n"
        "*1️⃣ Overnight Gap (0–3 pts)*\n"
        "Formula: `(open - prev_close) / prev_close`\n"
        "Cap-adjusted: large-cap ×2.5, mid-cap ×1.5.\n"
        "• 5%+ adjusted → 1 pt\n"
        "• 10%+ → 2 pts\n"
        "• 20%+ → 3 pts\n"
        "\n"
        "*2️⃣ Continuation (-1 to 2 pts)*\n"
        "Formula: `(last - open) / open`\n"
        "• Below -3% → -1 (fading)\n"
        "• 0% to +3% → +1 (holding)\n"
        "• +3%+ → +2 (extending)\n"
        "\n"
        "*3️⃣ 5-min Momentum (-1 to 1 pt)*\n"
        "Futu's built-in 5-min change field.\n"
        "• Below -1% → -1\n"
        "• +3%+ → +1\n"
    )

    part2 = (
        "*4️⃣ Price Strength (-1 to 1 pt)*\n"
        "Formula: `(last - low) / (high - low)`\n"
        "• Top 25% → +1\n"
        "• Bottom 25% → -1\n"
        "\n"
        "*5️⃣ Turnover Delta (0–3 pts)*\n"
        "New flow since last scan as % of yesterday's turnover.\n"
        "• 5%+ → 1 pt  •  15%+ → 2 pts  •  30%+ → 3 pts\n"
        "\n"
        "*6️⃣ Unusual Volume (0–3 pts)*\n"
        "Projected full-day turnover ÷ yesterday's turnover.\n"
        "• 1.5x → 1 pt  •  3x → 2 pts  •  5x → 3 pts\n"
        "\n"
        "*━━━ PATH B: INTRADAY BREAKOUT ━━━*\n"
        "Triggers when: open ≤ prev\\_close × 1.02 AND "
        "intraday move ≥ 10% from open.\n"
        "Same 6 signals scored, plus a *Breakout Bonus*.\n"
        "Same alert thresholds as gap-ups: 8+ alert, 4+ watch.\n"
        "\n"
        "*🚀 Breakout Bonus (0–2 pts)*\n"
        "Only for exceptional moves:\n"
        "• 20%+ from open → +2\n"
        "Signal 1 (gap) uses the intraday move instead "
        "of overnight gap for scoring.\n"
        "Min turnover HK$500k (filters penny noise).\n"
        "\n"
        "*⛔ Hard Veto*\n"
        "Caps score at 3 if:\n"
        "• Turnover delta < 3% AND today < 50% of yesterday's "
        "total turnover, OR\n"
        "• Today's total turnover < HK$50k, OR\n"
        "• Breakout with turnover < HK$500k\n"
        "\n"
        "*━━━ FOLLOW-UP TRACKING ━━━*\n"
        "All alerted & watch-tier stocks are auto-tracked.\n"
        "Every scan cycle re-checks tracked stocks:\n"
        "• 5%+ total gain since flagged AND still rising "
        "today → 📡 follow-up alert\n"
        "• Baseline resets after each follow-up (needs "
        "another 5% for next alert)\n"
        "• Tracked stocks auto-clean after 3 days\n"
        "Use /tracked to view all flagged stocks.\n"
        "\n"
        "*⏱ Cooldown:* 60 min between repeat alerts.\n"
        "\n"
        "*🕐 After Hours:* /scan returns End of Market Recap.\n"
        "\n"
        "*📨 Report:* Price, Change, Market Cap, Gap %, "
        "Continuation %, Momentum, Strength, Flow, Volume, Score /12"
    )

    await update.message.reply_text(part1, parse_mode="Markdown")
    await update.message.reply_text(part2, parse_mode="Markdown")


# ── /chat — LLM conversation mode ──────────────────────────────────────────────

CHAT_MSG = 0  # ConversationHandler state

CHAT_SYSTEM_PROMPT = """\
You are an expert Hong Kong stock market analyst assistant embedded in a \
Telegram trading bot called Odysseus.

You have deep knowledge of:
- HKEX listed companies, corporate actions (rights issues, placements, \
privatisations, takeovers, M&A, trading halts)
- Hong Kong securities regulations (Listing Rules, Takeovers Code, SFO)
- Technical analysis, market microstructure, CCASS shareholding patterns
- How to interpret HKEX filings and announcements

The bot you live in has these capabilities:
- /scan — scans the HK market for gap-up stocks and intraday movers, scores them
- /news — scans HKEX announcements for corporate actions (halts, takeovers, M&A, etc.)
- /corpactions — shows stocks on the Corporate Actions watchlist (auto-tracked for 90 days)
- /ipo — tracks recent IPO performance
- /joslist — runs a portfolio tracker that updates Google Sheets with price data
- /ccass — tracks CCASS shareholding changes (institutional flow)
- /monitor — toggles automatic scanning every 30 min during market hours

The bot automatically analyses M&A/Takeover/Privatisation deals when detected.
Stocks flagged from announcements are added to an "Extra Attention" watchlist \
and monitored for ≥3% intraday moves.

The user is a trader/investor monitoring HK stocks. Answer concisely and \
directly. Use bullet points where helpful. If you reference specific stocks, \
include the stock code.

Keep responses focused and actionable. Avoid disclaimers about not being \
financial advice unless explicitly asked.
"""


def _build_chat_context() -> str:
    """Build live bot state context for chat mode."""
    import json
    context_parts = []

    # Tracked movers
    tracked_file = Path(__file__).resolve().parent.parent / "data" / "tracked_movers.json"
    if tracked_file.exists():
        try:
            tracked = json.loads(tracked_file.read_text())
            if tracked:
                lines = ["Currently tracked scanner hits:"]
                for code, info in list(tracked.items())[:30]:  # limit to 30
                    name = info.get("name", "")
                    entry_price = info.get("entry_price", "")
                    date = info.get("date", "")[:10]
                    lines.append(f"- {code} {name}: entry ${entry_price} on {date}")
                context_parts.append("\n".join(lines))
        except Exception:
            pass

    # Attention watchlist
    attention = get_attention_stocks()
    if attention:
        lines = ["Corporate Actions watchlist (Extra Attention):"]
        for code, info in attention.items():
            key_info = info.get("key_info", "")
            lines.append(
                f"- {info['stock_code']} {info['stock_name']}: "
                f"{info['category']}"
                + (f" — {key_info[:100]}" if key_info else "")
            )
        context_parts.append("\n".join(lines))

    return "\n\n".join(context_parts) if context_parts else ""


async def cmd_chat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Enter LLM chat mode."""
    # Check API key
    api_key = os.environ.get("PERPLEXITY_API_KEY", "")
    if not api_key:
        await update.message.reply_text(
            "❌ PERPLEXITY_API_KEY not set. Cannot enter chat mode."
        )
        return ConversationHandler.END

    # Initialise conversation history for this user
    ctx.user_data["chat_history"] = [
        {"role": "system", "content": CHAT_SYSTEM_PROMPT},
    ]

    # Include live bot state as context
    live_ctx = _build_chat_context()
    if live_ctx:
        ctx.user_data["chat_history"].append(
            {"role": "system", "content": live_ctx}
        )

    await update.message.reply_text(
        "💬 *Chat Mode Active*\n\n"
        "Ask me anything about HK stocks, announcements, or market analysis.\n"
        "Type /done to exit chat mode.",
        parse_mode="Markdown",
    )
    return CHAT_MSG


async def chat_got_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Process a user message in chat mode."""
    from openai import OpenAI

    user_msg = update.message.text.strip()
    if not user_msg:
        return CHAT_MSG

    history = ctx.user_data.get("chat_history", [])
    history.append({"role": "user", "content": user_msg})

    # Keep history manageable (last 20 messages + system)
    system_msgs = [m for m in history if m["role"] == "system"]
    conv_msgs = [m for m in history if m["role"] != "system"]
    if len(conv_msgs) > 20:
        conv_msgs = conv_msgs[-20:]
    messages = system_msgs + conv_msgs

    await update.message.reply_text("🤔 Thinking…")

    try:
        client = OpenAI(
            api_key=os.environ.get("PERPLEXITY_API_KEY", ""),
            base_url="https://api.perplexity.ai",
        )

        response = client.chat.completions.create(
            model="sonar",
            messages=messages,
            temperature=0.3,
            max_tokens=1500,
        )

        reply = response.choices[0].message.content.strip()
        history.append({"role": "assistant", "content": reply})
        ctx.user_data["chat_history"] = history

        # Split long messages (Telegram 4096 char limit)
        for i in range(0, len(reply), 4000):
            chunk = reply[i:i + 4000]
            await update.message.reply_text(
                chunk, parse_mode="Markdown",
                disable_web_page_preview=True,
            )

    except Exception as e:
        logger.warning("Chat LLM call failed: %s", e)
        # Retry without markdown in case of parse error
        try:
            await update.message.reply_text(f"Reply:\n{reply}")
        except Exception:
            await update.message.reply_text(f"❌ Error: {e}")

    return CHAT_MSG


async def chat_done(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Exit chat mode."""
    ctx.user_data.pop("chat_history", None)
    await update.message.reply_text("👋 Chat mode ended.")
    return ConversationHandler.END


async def chat_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Cancel chat mode."""
    ctx.user_data.pop("chat_history", None)
    await update.message.reply_text("Chat cancelled.")
    return ConversationHandler.END


# ── Extra Attention stocks monitoring ──────────────────────────────────────────

# Track last alert time per attention stock to avoid spam (10-min cooldown)
_attention_last_alert: dict[str, datetime] = {}
ATTENTION_COOLDOWN_S = 600  # 10 minutes


async def _check_attention_stocks(ctx, analyzer):
    """Check extra attention stocks for significant movement and alert."""
    target_chat = CHAT_ID
    if not target_chat:
        return

    attention = get_attention_stocks()
    if not attention:
        return

    codes = list(attention.keys())
    all_snaps = []
    for i in range(0, len(codes), 400):
        batch = codes[i:i + 400]
        try:
            ret, snap = analyzer.quote_ctx.get_market_snapshot(batch)
            if ret == 0 and snap is not None and not snap.empty:
                all_snaps.append(snap)
        except Exception as e:
            logger.debug("Attention stock snapshot failed: %s", e)

    if not all_snaps:
        return

    import pandas as pd
    snapshot = pd.concat(all_snaps, ignore_index=True)

    alerts = []
    now = datetime.now()

    for _, row in snapshot.iterrows():
        code = row["code"]
        if code not in attention:
            continue

        info = attention[code]
        cur = float(row.get("last_price", 0) or 0)
        prev_close = float(row.get("prev_close_price", 0) or 0)
        if prev_close <= 0 or cur <= 0:
            continue

        daily_chg = (cur - prev_close) / prev_close * 100

        # Alert if moved ≥3% intraday
        if abs(daily_chg) < 3.0:
            continue

        # Cooldown check
        last = _attention_last_alert.get(code)
        if last and (now - last).total_seconds() < ATTENTION_COOLDOWN_S:
            continue

        _attention_last_alert[code] = now

        direction = "📈" if daily_chg > 0 else "📉"
        annotation = get_annotation(code) or info.get("category", "")
        turnover = float(row.get("turnover", 0) or 0)
        turnover_str = f"HK${turnover/1e6:.1f}M" if turnover >= 1e6 else f"HK${turnover:,.0f}"

        alerts.append(
            f"{direction} `{info['stock_code']}` *{info['stock_name']}*\n"
            f"  Price: ${cur:.3f}  ({daily_chg:+.1f}%)\n"
            f"  Turnover: {turnover_str}\n"
            f"  📢 {annotation}"
        )

    if alerts:
        header = "🔔 *Extra Attention — Stock Movement*\n"
        text = header + "\n" + "\n\n".join(alerts)
        await ctx.bot.send_message(
            chat_id=target_chat, text=text,
            parse_mode="Markdown",
        )


async def cmd_corp_actions(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Show all stocks on the Corporate Actions watchlist + sync to Google Sheets."""
    attention = get_attention_stocks()
    if not attention:
        await update.message.reply_text("No corporate actions being tracked.")
        return

    lines = ["*Corporate Actions*", ""]
    emojis = {
        "Trading Halt": "🔴", "Trading Resumption": "🟢",
        "Rights Issue": "💰", "Share Placement": "📊",
        "Privatisation": "🏛️", "Takeover": "⚠️", "M&A": "🔵",
    }

    for code, info in sorted(attention.items()):
        cat = info.get("category", "")
        emoji = emojis.get(cat, "📢")
        added = info.get("added", "")[:10]
        expires = info.get("expires", "")[:10]
        key_info = info.get("key_info", "")

        link = info.get("link", "")
        lines.append(f"{emoji} `{info['stock_code']}` *{info['stock_name']}*")
        lines.append(f"  Type: {cat}")
        if key_info:
            lines.append(f"  💡 _{key_info[:80]}_")
        lines.append(f"  Added: {added} → Expires: {expires}")
        if link:
            lines.append(f"  [Announcement]({link})")
        lines.append("")

    lines.append(f"_Total: {len(attention)} stocks tracked for 3 months_")

    # Sync to Google Sheets
    try:
        sheet_rows = [
            {"Code": code, "Name": info.get("stock_name", ""),
             "Category": info.get("category", ""),
             "Key Info": info.get("key_info", "")[:120],
             "Added": info.get("added", "")[:10],
             "Expires": info.get("expires", "")[:10],
             "Announcement": info.get("link", "")}
            for code, info in sorted(attention.items())
        ]
        desc = "Stocks flagged from HKEX announcements. Auto-added by /news. Retained 90 days."
        url = write_corporate_actions(sheet_rows, desc)
        lines.append(f"\n[Open in Google Sheets]({url})")
    except Exception as e:
        logger.warning("Google Sheets sync failed: %s", e)

    await update.message.reply_text(
        "\n".join(lines), parse_mode="Markdown",
        disable_web_page_preview=True,
    )


# ── Scheduled background scan ──────────────────────────────────────────────────

async def scheduled_scan(ctx: ContextTypes.DEFAULT_TYPE):
    """Called by the job queue — only acts when monitoring is ON."""
    if not monitoring_active:
        return

    global last_scan_time, _eod_saved_date
    hm, now_hkt = _hk_now()

    if not _hk_market_open():
        # Save EOD turnover once after market close (for unusual-volume signal)
        today_date = now_hkt.date()
        if hm >= 1600 and _eod_saved_date != today_date:
            try:
                a = get_analyzer()
                snapshot = a._fetch_full_snapshot(MARKET)
                if not snapshot.empty:
                    save_prev_day_turnover(MARKET, snapshot)
                    _eod_saved_date = today_date
                    logger.info("EOD turnover saved for %s", MARKET)
            except Exception as e:
                logger.exception("EOD save failed")
        return

    try:
        a = get_analyzer()

        # ── Run both scanners (raw results, no Telegram sends) ──
        intra_alerts, intra_watches = a.scan_intraday_movers(market=MARKET)
        gap_alerts, gap_watches = a.scan_gap_ups(market=MARKET)

        # Deduplicate: intraday takes priority
        seen = set()
        all_alerts, all_watches = [], []
        for row, result in intra_alerts:
            seen.add(result["code"])
            all_alerts.append((row, result))
        for row, result in gap_alerts:
            if result["code"] not in seen:
                seen.add(result["code"])
                all_alerts.append((row, result))
        for row, result in intra_watches:
            if result["code"] not in seen:
                seen.add(result["code"])
                all_watches.append((row, result))
        for row, result in gap_watches:
            if result["code"] not in seen:
                seen.add(result["code"])
                all_watches.append((row, result))

        # Track new stocks
        for row, result in all_alerts + all_watches:
            code = result["code"]
            tracked = a.load_tracked_stocks()
            if code not in tracked:
                a.add_to_tracking(code, str(row.get("name", "")),
                                  float(row.get("last_price", 0) or 0),
                                  float(row.get("change_5min", 0) or 0), MARKET)

        last_scan_time = datetime.now()

        target_chat = CHAT_ID

        # Send top 2 alerts to Telegram
        if all_alerts and target_chat:
            top_alerts = sorted(all_alerts, key=lambda x: x[1].get("score", 0), reverse=True)[:2]
            for row, result in top_alerts:
                trigger = result.get("trigger", "gap_up")
                if trigger == "gap_up":
                    msg = build_telegram_message(row, result)
                else:
                    msg = build_intraday_message(row, result)
                await ctx.bot.send_message(chat_id=target_chat, text=msg,
                                           parse_mode="Markdown")

        # Sync all results to Google Sheets
        try:
            result_info = _sync_scanner_hits_to_sheets(a)
            if result_info and target_chat:
                url, sheet_rows = result_info
                await ctx.bot.send_message(
                    chat_id=target_chat,
                    text=f"🤖 *Auto-Scan* — {len(all_alerts)} alerts, {len(all_watches)} watch\n"
                         f"[Full results in Google Sheets]({url})",
                    parse_mode="Markdown",
                    disable_web_page_preview=True,
                )
        except Exception as e:
            logger.warning("Auto-scan Sheets sync failed: %s", e)

        logger.info("Scheduled scan: %d alerts, %d watches",
                    len(all_alerts), len(all_watches))

        # ── Check Extra Attention stocks for movement ──
        await _check_attention_stocks(ctx, a)

    except Exception as e:
        logger.exception("Scheduled scan failed")


# ── Entrypoint ──────────────────────────────────────────────────────────────────

def main():
    if not BOT_TOKEN:
        print("Set TELEGRAM_BOT_TOKEN environment variable first.")
        print("  export TELEGRAM_BOT_TOKEN='your-token'")
        print("  export TELEGRAM_CHAT_ID='your-chat-id'")
        return

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("scan", cmd_scan))
    app.add_handler(CommandHandler("joslist", cmd_joslist))
    app.add_handler(CommandHandler("joadd", cmd_joadd))
    app.add_handler(CommandHandler("joremove", cmd_joremove))
    app.add_handler(CommandHandler("monitor", cmd_monitor))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("algo", cmd_algo))
    app.add_handler(CommandHandler("start", cmd_help))
    app.add_handler(CommandHandler("news", cmd_news))
    app.add_handler(CommandHandler("corpactions", cmd_corp_actions))
    app.add_handler(CommandHandler("ipo", cmd_ipo))
    app.add_handler(CallbackQueryHandler(deal_analysis_callback, pattern=r"^deal\|"))
    app.add_handler(CallbackQueryHandler(news_callback, pattern=r"^news\|"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, news_date_handler), group=1)

    # Debug conversation handler
    debug_conv = ConversationHandler(
        entry_points=[CommandHandler("debug", cmd_debug)],
        states={
            DEBUG_CODE: [MessageHandler(filters.TEXT & ~filters.COMMAND, debug_got_code)],
        },
        fallbacks=[CommandHandler("cancel", debug_cancel)],
    )
    app.add_handler(debug_conv)

    # CCASS conversation handler
    ccass_conv = ConversationHandler(
        entry_points=[CommandHandler("ccass", cmd_ccass)],
        states={
            CCASS_CODE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccass_got_code)],
            CCASS_SOURCE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccass_got_source)],
            CCASS_PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccass_got_period)],
            CCASS_INTERVAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccass_got_interval)],
        },
        fallbacks=[CommandHandler("cancel", ccass_cancel)],
    )
    app.add_handler(ccass_conv)

    # Chat conversation handler
    chat_conv = ConversationHandler(
        entry_points=[CommandHandler("chat", cmd_chat)],
        states={
            CHAT_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, chat_got_message)],
        },
        fallbacks=[
            CommandHandler("done", chat_done),
            CommandHandler("cancel", chat_cancel),
        ],
    )
    app.add_handler(chat_conv)

    # ── Schedule all recurring / daily jobs ──
    hkt = timezone(timedelta(hours=8))

    # First market scan at 9:15 AM HKT, then repeat every SCAN_INTERVAL_S
    app.job_queue.run_daily(
        scheduled_scan,
        time=dt_time(hour=9, minute=15, tzinfo=hkt),
        name="scan_morning",
    )
    app.job_queue.run_repeating(
        scheduled_scan,
        interval=SCAN_INTERVAL_S,
        first=dt_time(hour=9, minute=20, tzinfo=hkt),
    )

    # Schedule HKEX announcement checks at 9:15 AM and 4:15 PM HKT daily
    async def scheduled_news_check(context: ContextTypes.DEFAULT_TYPE):
        """Fetch HKEX announcements, auto-analyse deals, sync to Sheets."""
        news_chat = CHAT_ID
        if not news_chat:
            return
        try:
            results = scan_announcements(since_hours=18.0)
            if results:
                text = format_alerts(results)
                # Split long messages (Telegram 4096 char limit)
                for i in range(0, len(text), 4000):
                    chunk = text[i:i + 4000]
                    await context.bot.send_message(
                        chat_id=news_chat, text=chunk,
                        parse_mode="Markdown",
                        disable_web_page_preview=True,
                    )

                # Auto-add to Extra Attention watchlist
                added = add_from_scan_results(results)
                if added:
                    await context.bot.send_message(
                        chat_id=news_chat,
                        text=f"📌 Added {len(added)} stock(s) to Extra Attention watchlist (3-month tracking).",
                    )

                # Auto-run deal analysis for Privatisation/Takeover/M&A
                deal_items = [
                    r for r in results
                    if r.get("is_original", True)
                    and r.get("category") in ("Privatisation", "Takeover", "M&A")
                ]
                for item in deal_items:
                    code = item.get("stock_code", "")
                    name = item.get("stock_name", "")
                    cat = item.get("category", "")
                    try:
                        await context.bot.send_message(
                            chat_id=news_chat,
                            text=f"🔎 Auto-analysing deal: `{code}` *{name}* ({cat})…",
                            parse_mode="Markdown",
                        )
                        analysis = analyze_deal(
                            title=item.get("title", ""),
                            stock_code=code,
                            stock_name=name,
                            link=item.get("link", ""),
                        )
                        if analysis:
                            for i in range(0, len(analysis), 4000):
                                await context.bot.send_message(
                                    chat_id=news_chat, text=analysis[i:i + 4000],
                                    parse_mode="Markdown",
                                    disable_web_page_preview=True,
                                )
                        else:
                            await context.bot.send_message(
                                chat_id=news_chat,
                                text=f"❌ Deal analysis failed for {code} {name}.",
                            )
                    except Exception as e:
                        logger.warning("Auto deal analysis failed for %s: %s", code, e)

                logger.info("Announcement alert sent: %d items", len(results))
            else:
                logger.info("Announcement scan: nothing new")

            # Auto-sync Corporate Actions watchlist to Google Sheets
            try:
                attention = get_attention_stocks()
                if attention:
                    sheet_rows = [
                        {"Code": code, "Name": info.get("stock_name", ""),
                         "Category": info.get("category", ""),
                         "Key Info": info.get("key_info", "")[:120],
                         "Added": info.get("added", "")[:10],
                         "Expires": info.get("expires", "")[:10],
                         "Announcement": info.get("link", "")}
                        for code, info in sorted(attention.items())
                    ]
                    write_corporate_actions(sheet_rows,
                        "Stocks flagged from HKEX announcements. Auto-synced. Retained 90 days.")
                    logger.info("Corporate Actions auto-synced to Google Sheets: %d rows", len(sheet_rows))
            except Exception as e:
                logger.warning("Corporate Actions auto-sync to Sheets failed: %s", e)

        except Exception as e:
            logger.exception("Scheduled announcement check failed")

    async def scheduled_afternoon_check(context: ContextTypes.DEFAULT_TYPE):
        """Afternoon: announcement check + IPO scan to Google Sheets."""
        # Run the standard announcement check first
        await scheduled_news_check(context)

        # Then scan IPOs and sync to Google Sheets
        news_chat = CHAT_ID
        try:
            a = get_analyzer()
            rows = fetch_recent_ipos(a.quote_ctx, market=MARKET)
            if rows:
                description = (
                    "--- IPO TRACKER ---",
                    "Recent HK IPOs tracked for 90 days after listing.",
                    "Data: Futu 'Recent IPOs' plate + market snapshot + industry plates.",
                    "IPO Price = official offer price if available, otherwise first-day opening price.",
                    f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')} HKT",
                )
                url = write_ipo_tracker(rows, "\n".join(description))

                with_return = [r for r in rows if isinstance(r.get("Return Since IPO %"), (int, float))]
                n_up = sum(1 for r in with_return if r["Return Since IPO %"] > 0)
                n_down = sum(1 for r in with_return if r["Return Since IPO %"] < 0)

                if news_chat:
                    await context.bot.send_message(
                        chat_id=news_chat,
                        text=f"📋 *IPO Tracker updated* — {len(rows)} recent IPOs\n"
                             f"📈 {n_up} up  📉 {n_down} down since listing\n\n"
                             f"[Open in Google Sheets]({url})",
                        parse_mode="Markdown",
                        disable_web_page_preview=True,
                    )
                logger.info("IPO tracker auto-synced: %d IPOs", len(rows))
            else:
                logger.info("IPO scan: no recent IPOs found")
        except Exception as e:
            logger.warning("Scheduled IPO scan failed: %s", e)

    app.job_queue.run_daily(
        scheduled_news_check,
        time=dt_time(hour=9, minute=15, tzinfo=hkt),
        name="news_morning",
    )
    app.job_queue.run_daily(
        scheduled_afternoon_check,
        time=dt_time(hour=16, minute=15, tzinfo=hkt),
        name="news_afternoon",
    )

    # Pre-fetch yesterday's turnover for all stocks via yfinance
    # Runs in a background thread so the bot starts immediately.
    def _bg_prefetch():
        try:
            prefetch_yesterday_turnover(MARKET)
        except Exception as e:
            logger.warning("yfinance prefetch failed (signals 5/6 will use fallback): %s", e)

    threading.Thread(target=_bg_prefetch, daemon=True).start()

    logger.info("Bot started — market=%s, interval=%ds", MARKET, SCAN_INTERVAL_S)
    app.run_polling()


if __name__ == "__main__":
    main()
